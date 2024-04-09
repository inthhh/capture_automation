from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image
from datetime import datetime
import os
import pandas as pd
from db import execute
from PIL import Image as PILImage
from io import BytesIO

data_list = []
before_location = ''
before_area = ''
before_title = ''

def excel_start(tab_title):
    exl_wb = Workbook()
    exl_ws = exl_wb.active
    exl_ws.title = tab_title

    exl_header = ['Location', 'Area', 'Title', 'Description', 'Contents', 'Check', 'Guide', 'Image']
    exl_ws.append(exl_header)

    exl_ws.column_dimensions['A'].width = 15
    exl_ws.column_dimensions['B'].width = 40
    exl_ws.column_dimensions['C'].width = 15
    exl_ws.column_dimensions['D'].width = 15
    exl_ws.column_dimensions['E'].width = 50
    exl_ws.column_dimensions['F'].width = 5
    exl_ws.column_dimensions['G'].width = 15
    exl_ws.column_dimensions['H'].width = 26
    exl_ws.sheet_view.showGridLines = False
    exl_ws.freeze_panes = 'A2'

    return exl_ws, exl_wb

def excel_end(exl_ws, exl_wb, start_time, url, img_html):
    row_num = exl_ws.max_row + 1
    for i in range(1, 9) :
        exl_ws.cell(row=row_num, column=i).border = Border(top=Side(style='thin', color='555555'))

    # df = pd.DataFrame(data_list)
    file_name_base = url.replace('https://www.samsung.com/', '').replace('/', '-')
    file_name_base = file_name_base.strip('-')
    current_time = datetime.now().strftime("%Y%m%d%H%M%S")
    end_time = datetime.now()

    path = "./result/" + end_time.strftime("%Y%m%d")
    os.makedirs(path, exist_ok=True)
    final_file_name = f"{path}/{file_name_base}_{current_time}.xlsx"
    exl_wb.save(final_file_name)

    time_elapsed = end_time - start_time
    print("FINISH", url, 'at : ', end_time.strftime("%Y/%m/%d %H:%M:%S"), "(", time_elapsed, ")")
    print(f"SAVE AT : {final_file_name}")
    img_html += ("<div class=\"row\">\n"
                 "<div class=\"col12\">\n"
                 "<button type=\"button\" class=\"btn btn-primary btn-lg\" id=\"sendButton\">Send</button>\n"
                 "</div>\n"
                 "</div>\n"
                 "</div>\n"
                 "<script>\n"
                 "document.getElementById('sendButton').addEventListener('click', function() {\n"
                 "let message = '';\n"
                 "const checkboxes = document.querySelectorAll('.btn-check');\n"
                 "checkboxes.forEach((checkbox, index) => {\n"
                 "message += (index + 1) + '번 ' + (checkbox.checked ? 'Yes' : 'No') + '\\n';});alert(message);});</script>\n"
                 "</body>\n</html>")

    html_file_name = f"{path}/html/{file_name_base}_{current_time}.html"
    with open(html_file_name, "w", encoding="utf-8") as file :
        file.write(img_html)



def excel_save_data(exl_ws, col_location, col_area, col_title, cell_description, cell_contents, cell_check, cell_inspection, img_file, raw_data_meta, dev_status):

    global data_list, before_location, before_area, before_title

    tab_menu_changed = col_location != before_location

    cell_location = '' if col_location == before_location else col_location
    cell_area = col_area if tab_menu_changed or col_area != before_area else ''
    cell_title = '' if col_title == before_title else col_title

    saved_item = [cell_location, cell_area, cell_title, cell_description, cell_contents, cell_check, cell_inspection]
    print(col_location, col_area, col_title, cell_description, cell_check, cell_inspection)
    exl_ws.append(saved_item)

    row_num = exl_ws.max_row

    if cell_location :
        exl_ws.cell(row=row_num, column=1).border = Border(top=Side(style='thin', color='555555'), right=Side(style='thin', color='555555'))
    else :
        exl_ws.cell(row=row_num, column=1).border = Border(right=Side(style='thin', color='555555'))
    exl_ws.cell(row=row_num, column=1).alignment = Alignment(vertical='center')

    if cell_area :
        exl_ws.cell(row=row_num, column=2).border = Border(top=Side(style='thin', color='555555'), right=Side(style='thin', color='555555'))
    else :
        exl_ws.cell(row=row_num, column=2).border = Border(right=Side(style='thin', color='555555'))
    exl_ws.cell(row=row_num, column=2).alignment = Alignment(vertical='center')

    if cell_title :
        exl_ws.cell(row=row_num, column=3).border = Border(top=Side(style='thin', color='555555'), right=Side(style='thin', color='555555'))
    else :
        exl_ws.cell(row=row_num, column=3).border = Border(right=Side(style='thin', color='555555'))
    exl_ws.cell(row=row_num, column=3).alignment = Alignment(vertical='center')

    exl_ws.cell(row=row_num, column=4).border = Border(top=Side(style='thin', color='555555'), right=Side(style='thin', color='555555'))
    exl_ws.cell(row=row_num, column=4).alignment = Alignment(vertical='center')
    exl_ws.cell(row=row_num, column=5).border = Border(top=Side(style='thin', color='555555'), right=Side(style='thin', color='555555'))
    exl_ws.cell(row=row_num, column=5).alignment = Alignment(vertical='center', wrap_text=True)
    exl_ws.cell(row=row_num, column=6).border = Border(top=Side(style='thin', color='555555'), right=Side(style='thin', color='555555'))
    exl_ws.cell(row=row_num, column=6).alignment = Alignment(vertical='center', horizontal='center')
    exl_ws.cell(row=row_num, column=7).border = Border(top=Side(style='thin', color='555555'), right=Side(style='thin', color='555555'))
    exl_ws.cell(row=row_num, column=7).alignment = Alignment(vertical='center', wrap_text=True)
    exl_ws.cell(row=row_num, column=8).border = Border(top=Side(style='thin', color='555555'), right=Side(style='thin', color='555555'))

    if img_file != '':
        original_width, original_height = img_file.size

        scale_factor = 200 / original_width
        new_height = int(original_height * scale_factor)

        pil_img_resized = img_file.resize((200, new_height), PILImage.LANCZOS)

        image_bytes_resized = BytesIO()
        pil_img_resized.save(image_bytes_resized, format=img_file.format)
        image_bytes_resized.seek(0)
        img = Image(image_bytes_resized)
        img.anchor = f'H{row_num}'
        exl_ws.row_dimensions[row_num].height = new_height * 0.75 + 1
        exl_ws.add_image(img, f'H{row_num}')

    before_location = col_location
    before_area = col_area
    before_title = col_title

    if dev_status == 'dev':
        insert_query = "INSERT INTO qa_result (date,rhq, subsidiary, site_code, page_type, category, location, area, title, description, contents, check_result, check_reason) values (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);"
        execute(insert_query, (
            raw_data_meta["date"], raw_data_meta["rhq"], raw_data_meta["subsidiary"], raw_data_meta["site_code"],
            raw_data_meta["page_type"],raw_data_meta["category"], col_location, col_area, col_title, cell_description, cell_contents, cell_check,
            cell_inspection))
    date_now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(date_now, ": ", col_location, col_area, col_title, cell_contents, cell_check, cell_inspection)



