import requests
import csv

from bs4 import BeautifulSoup
from datetime import datetime
# from pandas import ExcelWriter
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
import pandas as pd
from process_components import *
from data_management import *
from datetime import datetime
from region import region

region = region
# url_codes = ['uk', 'fr', 'de', 'it', 'es', 'sg', 'id', 'ph', 'au', 'my', 'th', 'vn', 'ar', 'mx', 'br', 'in', 'sa', 'sa_en', 'ca', 'ca_fr', 'ae', 'ae_ar']


def qa (url_code):
    raw_data_meta = {
        "date": datetime.today().strftime('%Y-%m-%d'),
        "rhq": region[url_code]["rhq"],
        "subsidiary": region[url_code]["subsidiary"],
        "site_code": url_code,
        "page_type": "Home",
        "category": "Home",
    }
    url = 'https://www.samsung.com/' + url_code + '/'
    sitecode = url.split('/')[3]
    response = requests.get(url)
    html_all = BeautifulSoup(response.text, 'lxml')
    start_time = datetime.now()
    print("START", url, 'at : ', start_time.strftime("%Y/%m/%d %H:%M:%S"))
    try:
        content_all = html_all.select(
            '#content > div > div > div.responsivegrid.aem-GridColumn.aem-GridColumn--default--12 > div > div')

        exl_wb = Workbook()
        exl_ws = exl_wb.active
        exl_ws.title = "Home"

        exl_header = ['Location', 'Area', 'Title', 'Description', 'Contents', 'Check', 'Guide','Image']
        exl_ws.append(exl_header)

        exl_ws.column_dimensions['A'].width = 15
        exl_ws.column_dimensions['B'].width = 40
        exl_ws.column_dimensions['C'].width = 15
        exl_ws.column_dimensions['D'].width = 15
        exl_ws.column_dimensions['E'].width = 50
        exl_ws.column_dimensions['F'].width = 5
        exl_ws.column_dimensions['G'].width = 15
        exl_ws.column_dimensions['H'].width = 26
        exl_ws.sheet_view.showGridLines = False
        exl_ws.freeze_panes = 'A2'

        no_content_comp_div = 0
        for content_comp_div in content_all:
            no_content_comp_div += 1
            content_comp_name = content_comp_div.get('class', [])
            if 'ho-g-home-kv-carousel' in content_comp_name:

                cell_value = process_class_attributes(exl_ws, content_comp_div,
                                                      "", "", "", "",
                                                      "N", "section",
                                                      {"home-kv-carousel--width-large": "1920px"}, "1440px", "N", raw_data_meta)
                if cell_value == '1920px':
                    bg_image_desktop_width = 1920
                    bg_image_mobile_width = 720
                else:
                    bg_image_desktop_width = 1440
                    bg_image_mobile_width = 720

                cell_value = process_class_attributes(exl_ws, content_comp_div,
                                                      "", "", "", "",
                                                      "N", "section",
                                                      {"home-kv-carousel--height-large": "Desktop:810px / Mobile:640px",
                                                          "home-kv-carousel--height-medium": "Desktop:640px / Mobile:540px",
                                                          "home-kv-carousel--height-smedium": "Desktop:344px / Mobile:400px",
                                                          "home-kv-carousel--height-small": "Desktop:320px / Mobile:320px"},
                                                      "N", "N", raw_data_meta)
                if cell_value == 'Desktop:810px / Mobile:640px':
                    bg_image_desktop_height = 810
                    bg_image_mobile_height = 1280
                elif cell_value == 'Desktop:640px / Mobile:540px':
                    bg_image_desktop_height = 640
                    bg_image_mobile_height = 1080
                elif cell_value == 'Desktop:344px / Mobile:400px':
                    bg_image_desktop_height = 344
                    bg_image_mobile_height = 800
                elif cell_value == 'Desktop:320px / Mobile:320px':
                    bg_image_desktop_height = 320
                    bg_image_mobile_height = 640

                carousel_no = 0
                for content_carousel_div in content_comp_div.select(
                        'section > div.home-kv-carousel__container > div.home-kv-carousel__wrapper > div'):
                    carousel_no += 1

                    cell_value = process_attributes_value(exl_ws, content_carousel_div,
                                                          "KeyVisual", "KV" + str(carousel_no), "Headline Text",
                                                          'Desktop',
                                                          "div > div > div.home-kv-carousel__text-wrap", "H",
                                                          "home-kv-carousel__headline",
                                                          "data-desktop-headline-text", "Y", "Y", raw_data_meta)
                    cell_value = process_attributes_value(exl_ws, content_carousel_div,
                                                          "KeyVisual", "KV" + str(carousel_no), "Headline Text",
                                                          'Mobile',
                                                          "div > div > div.home-kv-carousel__text-wrap", "H",
                                                          "home-kv-carousel__headline",
                                                          "data-mobile-headline-text", "Y", "Y", raw_data_meta)
                    cell_value = process_attributes_value(exl_ws, content_carousel_div,
                                                          "KeyVisual", "KV" + str(carousel_no), "Description Text",
                                                          'Desktop',
                                                          "div > div > div.home-kv-carousel__text-wrap", "p",
                                                          "home-kv-carousel__desc",
                                                          "data-desktop-description", "N", "Y", raw_data_meta)
                    cell_value = process_attributes_value(exl_ws, content_carousel_div,
                                                          "KeyVisual", "KV" + str(carousel_no), "Description Text",
                                                          'Mobile',
                                                          "div > div > div.home-kv-carousel__text-wrap", "p",
                                                          "home-kv-carousel__desc",
                                                          "data-mobile-description", "N", "Y", raw_data_meta)

                    process_cta_buttons(exl_ws, content_carousel_div, "KeyVisual", "KV" + str(carousel_no), "CTA",
                                        "div > div > div.home-kv-carousel__text-wrap > div.home-kv-carousel__cta-wrap", raw_data_meta)

                    process_background_image(exl_ws, content_carousel_div, "KeyVisual", "KV" + str(carousel_no),
                                             "BG Image",
                                             "div > div > div.home-kv-carousel__background-media-wrap",
                                             bg_image_desktop_width, bg_image_desktop_height, bg_image_mobile_width,
                                             bg_image_mobile_height, "N", "Y", raw_data_meta)
            if 'cm-g-text-block-container' in content_comp_name:

                if content_comp_div.select('div > div.ho-g-showcase-card-tab'):
                    col_location = process_label_text(exl_ws, content_comp_div, "", "", "", "",
                                                      "section > div > div", "H", "text-block-container__headline",
                                                      "Y", "N", raw_data_meta)
                    for content_comp_co02 in content_comp_div.select('div.showcase-card-tab'):

                        tab_no = 0
                        for comp_co02_tab in content_comp_co02.select('div > ul.tab__list > li'):
                            tab_no += 1
                            cell_area = process_label_text(exl_ws, comp_co02_tab, "", "", "", "",
                                                           "N", "button", "tab__item-title", "N", "N", raw_data_meta)

                            comp_co02_layout = content_comp_co02.select(
                                'div.showcase-card-tab__card-wrap > div > div.showcase-card-tab__card-items')[
                                tab_no - 1]

                            card_layout_option = ''
                            card_no = 0

                            for card_compo in comp_co02_layout.find_all("div", class_='showcase-card-tab-card'):
                                if ('showcase-card-tab-card--large' in card_compo.attrs['class']
                                        or 'showcase-card-tab-card--product-large' in card_compo.attrs['class']
                                        or 'showcase-card-tab-card--vertical' in card_compo.attrs['class']
                                        or 'showcase-card-tab-card--product-vertical' in card_compo.attrs['class']):
                                    card_layout_option += 'L'
                                elif ('showcase-card-tab-card--small' in card_compo.attrs['class']
                                      or 'showcase-card-tab-card--product-small' in card_compo.attrs['class']):
                                    card_layout_option += 'S'

                            options = {'SSSSSSSS': ['8 Card',
                                                    ['Top Left', 'Bottom Left', 'Top Center Left',
                                                     'Bottom Center Left',
                                                     'Top Center Right', 'Bottom Center Right', 'Top Right',
                                                     'Bottom Right']],
                                       'LSSSSSS': ['7 Card (1-2-2-2)',
                                                   ['Left', 'Top Center Left', 'Bottom Center Left',
                                                    'Top Center Right',
                                                    'Bottom Center Right', 'Top Right', 'Bottom Right']],
                                       'SSLSSSS': ['7 Card (2-1-2-2)',
                                                   ['Top Left', 'Bottom Left', 'Center Left', 'Top Center Right',
                                                    'Bottom Center Right', 'Top Right', 'Bottom Right']],
                                       'SSSSLSS': ['7 Card (2-2-1-2)',
                                                   ['Top Left', 'Bottom Left', 'Top Center Left',
                                                    'Bottom Center Left',
                                                    'Center Right', 'Top Right', 'Bottom Right']],
                                       'SSSSSSL': ['7 Card (2-2-2-1)',
                                                   ['Top Left', 'Bottom Left', 'Top Center Left',
                                                    'Bottom Center Left',
                                                    'Top Center Right', 'Bottom Center Right', 'Right']],
                                       'SSSSSS': ['6 Card',
                                                  ['Top Left', 'Bottom Left', 'Top Center', 'Bottom Center',
                                                   'Top Right',
                                                   'Bottom Right']],
                                       'SSSSL': ['5 Card (2-2-1)',
                                                 ['Top Left', 'Bottom Left', 'Top Center', 'Bottom Center',
                                                  'Right']],
                                       'SSLSS': ['5 Card (2-1-2)',
                                                 ['Top Left', 'Bottom Left', 'Center', 'Top Right',
                                                  'Bottom Right']],
                                       'LSSSS': ['5 Card (1-2-2)',
                                                 ['Left', 'Top Center', 'Bottom Center', 'Top Right',
                                                  'Bottom Right']],
                                       'LLL': ['3 Card', ['Left', 'Center', 'Right']]}

                            for x in options.keys():
                                if x == card_layout_option:
                                    card_number = options[x][0]
                                    card_layout = options[x][1]

                            card_list_no = 0
                            for card_list in card_layout:
                                card_list_no += 1
                                print(card_list, card_list_no)
                                card_compo_detail = \
                                    comp_co02_layout.find_all("div", class_='showcase-card-tab-card')[card_list_no - 1]
                                card_type = process_class_attributes(exl_ws, card_compo_detail, "", "", "", "",
                                                                     'N', 'N',
                                                                     {"showcase-card-tab-card--large": "Full Breed",
                                                                      "showcase-card-tab-card--small": "Full Breed",
                                                                      "showcase-card-tab-card--vertical": "Full Breed",
                                                                      "showcase-card-tab-card--product-large": "Product",
                                                                      "showcase-card-tab-card--product-small": "Product",
                                                                      "showcase-card-tab-card--product-vertical": "Product"},
                                                                     'N',
                                                                     "N", raw_data_meta)

                                if card_type == 'Full Breed':
                                    if card_number == '3 Card':
                                        bg_image_desktop_width = 448
                                        bg_image_desktop_height = 684
                                        bg_image_mobile_width = 624
                                        bg_image_mobile_height = 352

                                    if (card_number == '5 Card (1-2-2)' and card_list == 'Left') or (
                                            card_number == '5 Card (2-1-2)' and card_list == 'Center') or (
                                            card_number == '5 Card (2-2-1)' and card_list == 'Right'):
                                        bg_image_desktop_width = 684
                                        bg_image_desktop_height = 684
                                        bg_image_mobile_width = 624
                                        bg_image_mobile_height = 352
                                    elif card_number == '5 Card (1-2-2)' or card_number == '5 Card (2-1-2)' or card_number == '5 Card (2-2-1)':
                                        bg_image_desktop_width = 330
                                        bg_image_desktop_height = 330
                                        bg_image_mobile_width = 296
                                        bg_image_mobile_height = 352

                                    if card_number == '6 Card':
                                        bg_image_desktop_width = 448
                                        bg_image_desktop_height = 330
                                        bg_image_mobile_width = 296
                                        bg_image_mobile_height = 352

                                    if (card_number == '7 Card (2-2-2-1)' and card_list == 'Right') or (
                                            card_number == '7 Card (2-2-1-2)' and card_list == 'Center Right') or (
                                            card_number == '7 Card (2-1-2-2)' and card_list == 'Center Left') or (
                                            card_number == '7 Card (1-2-2-2)' and card_list == 'Left'):
                                        bg_image_desktop_width = 330
                                        bg_image_desktop_height = 684
                                        bg_image_mobile_width = 624
                                        bg_image_mobile_height = 352
                                    elif card_number == '7 Card (2-2-2-1)' or card_number == '7 Card (2-2-1-2)' or card_number == '7 Card (2-1-2-2)' or card_number == '7 Card (1-2-2-2)':
                                        bg_image_desktop_width = 330
                                        bg_image_desktop_height = 330
                                        bg_image_mobile_width = 296
                                        bg_image_mobile_height = 352

                                    if card_number == '8 Card':
                                        bg_image_desktop_width = 330
                                        bg_image_desktop_height = 330
                                        bg_image_mobile_width = 296
                                        bg_image_mobile_height = 352

                                if card_type == 'Product':
                                    if card_number == '3 Card':
                                        bg_image_desktop_width = 376
                                        bg_image_desktop_height = 376
                                        bg_image_mobile_width = 264
                                        bg_image_mobile_height = 264

                                    if (card_number == '5 Card (1-2-2)' and card_list == 'Left') or (
                                            card_number == '5 Card (2-1-2)' and card_list == 'Center') or (
                                            card_number == '5 Card (2-2-1)' and card_list == 'Right'):
                                        bg_image_desktop_width = 376
                                        bg_image_desktop_height = 376
                                        bg_image_mobile_width = 264
                                        bg_image_mobile_height = 264
                                    elif card_number == '5 Card (1-2-2)' or card_number == '5 Card (2-1-2)' or card_number == '5 Card (2-2-1)':
                                        bg_image_desktop_width = 160
                                        bg_image_desktop_height = 160
                                        bg_image_mobile_width = 192
                                        bg_image_mobile_height = 192

                                    if card_number == '6 Card':
                                        bg_image_desktop_width = 160
                                        bg_image_desktop_height = 160
                                        bg_image_mobile_width = 192
                                        bg_image_mobile_height = 192

                                    if (card_number == '7 Card (2-2-2-1)' and card_list == 'Right') or (
                                            card_number == '7 Card (2-2-1-2)' and card_list == 'Center Right') or (
                                            card_number == '7 Card (2-1-2-2)' and card_list == 'Center Left') or (
                                            card_number == '7 Card (1-2-2-2)' and card_list == 'Left'):
                                        bg_image_desktop_width = 282
                                        bg_image_desktop_height = 282
                                        bg_image_mobile_width = 264
                                        bg_image_mobile_height = 264
                                    elif card_number == '7 Card (2-2-2-1)' or card_number == '7 Card (2-2-1-2)' or card_number == '7 Card (2-1-2-2)' or card_number == '7 Card (1-2-2-2)':
                                        bg_image_desktop_width = 160
                                        bg_image_desktop_height = 160
                                        bg_image_mobile_width = 192
                                        bg_image_mobile_height = 192

                                    if card_number == '8 Card':
                                        bg_image_desktop_width = 160
                                        bg_image_desktop_height = 160
                                        bg_image_mobile_width = 192
                                        bg_image_mobile_height = 192
                                if " " in card_list:
                                    card_size = "Small"
                                else:
                                    card_size = "Large"
                                cell_value = process_label_text(exl_ws, card_compo_detail, col_location,
                                                                cell_area + " | " + card_list + "(" + card_size + ")",
                                                                "Text", "Desktop",
                                                                "N", "span",
                                                                "showcase-card-tab-card__product-name--desktop",
                                                                "Y", "Y", raw_data_meta)
                                cell_value = process_label_text(exl_ws, card_compo_detail, col_location,
                                                                cell_area + " | " + card_list + "(" + card_size + ")",
                                                                "Text", "Mobile",
                                                                "N", "span",
                                                                "showcase-card-tab-card__product-name--mobile", "Y",
                                                                "Y", raw_data_meta)

                                cell_value = process_label_text(exl_ws, card_compo_detail, col_location,
                                                                cell_area + " | " + card_list + "(" + card_size + ")",
                                                                "Text",
                                                                "Desktop",
                                                                "N", "span",
                                                                "showcase-card-tab-card__product-description--desktop",
                                                                "N", "Y", raw_data_meta)
                                cell_value = process_label_text(exl_ws, card_compo_detail, col_location,
                                                                cell_area + " | " + card_list + "(" + card_size + ")",
                                                                "Text",
                                                                "Mobile",
                                                                "N", "span",
                                                                "showcase-card-tab-card__product-description--mobile",
                                                                "N", "Y", raw_data_meta)
                                process_label_text_cta(exl_ws, card_compo_detail, col_location,
                                                       cell_area + " | " + card_list + "(" + card_size + ")", "CTA",
                                                       "N", "span.cta", raw_data_meta)

                                process_background_image(exl_ws, card_compo_detail, col_location,
                                                         cell_area + " | " + card_list + "(" + card_size + ")",
                                                         "BG Image",
                                                         "div.showcase-card-tab-card__img-wrap",
                                                         bg_image_desktop_width, bg_image_desktop_height,
                                                         bg_image_mobile_width, bg_image_mobile_height, "Y", "Y", raw_data_meta)

            if 'ho-g-showcase-card-tab' in content_comp_name:

                col_location = 'Latest Deals'
                for content_comp_co02 in content_comp_div.select('div.showcase-card-tab'):

                    tab_no = 0
                    for comp_co02_tab in content_comp_co02.select('div > ul.tab__list > li'):
                        tab_no += 1
                        cell_area = process_label_text(exl_ws, comp_co02_tab, "", "", "", "",
                                                       "N", "button", "tab__item-title", "N", "N", raw_data_meta)

                        comp_co02_layout = content_comp_co02.select(
                            'div.showcase-card-tab__card-wrap > div > div.showcase-card-tab__card-items')[
                            tab_no - 1]

                        card_layout_option = ''
                        card_no = 0

                        for card_compo in comp_co02_layout.find_all("div", class_='showcase-card-tab-card'):
                            if 'showcase-card-tab-card--large' in card_compo.attrs[
                                'class'] or 'showcase-card-tab-card--product-large' in card_compo.attrs[
                                'class'] or 'showcase-card-tab-card--vertical' in card_compo.attrs[
                                'class'] or 'showcase-card-tab-card--product-vertical' in card_compo.attrs['class']:
                                card_layout_option += 'L'
                            elif 'showcase-card-tab-card--small' in card_compo.attrs[
                                'class'] or 'showcase-card-tab-card--product-small' in card_compo.attrs['class']:
                                card_layout_option += 'S'

                        options = {'SSSSSSSS': ['8 Card',
                                                ['Top Left', 'Bottom Left', 'Top Center Left', 'Bottom Center Left',
                                                 'Top Center Right', 'Bottom Center Right', 'Top Right',
                                                 'Bottom Right']],
                                   'LSSSSSS': ['7 Card (1-2-2-2)',
                                               ['Left', 'Top Center Left', 'Bottom Center Left', 'Top Center Right',
                                                'Bottom Center Right', 'Top Right', 'Bottom Right']],
                                   'SSLSSSS': ['7 Card (2-1-2-2)',
                                               ['Top Left', 'Bottom Left', 'Center Left', 'Top Center Right',
                                                'Bottom Center Right', 'Top Right', 'Bottom Right']],
                                   'SSSSLSS': ['7 Card (2-2-1-2)',
                                               ['Top Left', 'Bottom Left', 'Top Center Left', 'Bottom Center Left',
                                                'Center Right', 'Top Right', 'Bottom Right']],
                                   'SSSSSSL': ['7 Card (2-2-2-1)',
                                               ['Top Left', 'Bottom Left', 'Top Center Left', 'Bottom Center Left',
                                                'Top Center Right', 'Bottom Center Right', 'Right']],
                                   'SSSSSS': ['6 Card',
                                              ['Top Left', 'Bottom Left', 'Top Center', 'Bottom Center',
                                               'Top Right',
                                               'Bottom Right']],
                                   'SSSSL': ['5 Card (2-2-1)',
                                             ['Top Left', 'Bottom Left', 'Top Center', 'Bottom Center', 'Right']],
                                   'SSLSS': ['5 Card (2-1-2)',
                                             ['Top Left', 'Bottom Left', 'Center', 'Top Right', 'Bottom Right']],
                                   'LSSSS': ['5 Card (1-2-2)',
                                             ['Left', 'Top Center', 'Bottom Center', 'Top Right', 'Bottom Right']],
                                   'LLL': ['3 Card', ['Left', 'Center', 'Right']]}

                        for x in options.keys():
                            if x == card_layout_option:
                                card_number = options[x][0]
                                card_layout = options[x][1]

                        card_list_no = 0
                        for card_list in card_layout:
                            card_list_no += 1

                            card_compo_detail = comp_co02_layout.find_all("div", class_='showcase-card-tab-card')[
                                card_list_no - 1]
                            card_type = process_class_attributes(exl_ws, card_compo_detail, "", "", "", "",
                                                                 'N', 'N',
                                                                 {"showcase-card-tab-card--large": "Full Breed",
                                                                  "showcase-card-tab-card--small": "Full Breed",
                                                                  "showcase-card-tab-card--vertical": "Full Breed",
                                                                  "showcase-card-tab-card--product-large": "Product",
                                                                  "showcase-card-tab-card--product-small": "Product",
                                                                  "showcase-card-tab-card--product-vertical": "Product"},
                                                                 'N',
                                                                 "N", raw_data_meta)

                            if card_type == 'Full Breed':
                                if card_number == '3 Card':
                                    bg_image_desktop_width = 448
                                    bg_image_desktop_height = 684
                                    bg_image_mobile_width = 624
                                    bg_image_mobile_height = 352

                                if (card_number == '5 Card (1-2-2)' and card_list == 'Left') or (
                                        card_number == '5 Card (2-1-2)' and card_list == 'Center') or (
                                        card_number == '5 Card (2-2-1)' and card_list == 'Right'):
                                    bg_image_desktop_width = 684
                                    bg_image_desktop_height = 684
                                    bg_image_mobile_width = 624
                                    bg_image_mobile_height = 352
                                elif card_number == '5 Card (1-2-2)' or card_number == '5 Card (2-1-2)' or card_number == '5 Card (2-2-1)':
                                    bg_image_desktop_width = 330
                                    bg_image_desktop_height = 330
                                    bg_image_mobile_width = 296
                                    bg_image_mobile_height = 352

                                if card_number == '6 Card':
                                    bg_image_desktop_width = 448
                                    bg_image_desktop_height = 330
                                    bg_image_mobile_width = 296
                                    bg_image_mobile_height = 352

                                if (card_number == '7 Card (2-2-2-1)' and card_list == 'Right') or (
                                        card_number == '7 Card (2-2-1-2)' and card_list == 'Center Right') or (
                                        card_number == '7 Card (2-1-2-2)' and card_list == 'Center Left') or (
                                        card_number == '7 Card (1-2-2-2)' and card_list == 'Left'):
                                    bg_image_desktop_width = 330
                                    bg_image_desktop_height = 684
                                    bg_image_mobile_width = 624
                                    bg_image_mobile_height = 352
                                elif card_number == '7 Card (2-2-2-1)' or card_number == '7 Card (2-2-1-2)' or card_number == '7 Card (2-1-2-2)' or card_number == '7 Card (1-2-2-2)':
                                    bg_image_desktop_width = 330
                                    bg_image_desktop_height = 330
                                    bg_image_mobile_width = 296
                                    bg_image_mobile_height = 352

                                if card_number == '8 Card':
                                    bg_image_desktop_width = 330
                                    bg_image_desktop_height = 330
                                    bg_image_mobile_width = 296
                                    bg_image_mobile_height = 352

                            if card_type == 'Product':
                                if card_number == '3 Card':
                                    bg_image_desktop_width = 376
                                    bg_image_desktop_height = 376
                                    bg_image_mobile_width = 264
                                    bg_image_mobile_height = 264

                                if (card_number == '5 Card (1-2-2)' and card_list == 'Left') or (
                                        card_number == '5 Card (2-1-2)' and card_list == 'Center') or (
                                        card_number == '5 Card (2-2-1)' and card_list == 'Right'):
                                    bg_image_desktop_width = 376
                                    bg_image_desktop_height = 376
                                    bg_image_mobile_width = 264
                                    bg_image_mobile_height = 264
                                elif card_number == '5 Card (1-2-2)' or card_number == '5 Card (2-1-2)' or card_number == '5 Card (2-2-1)':
                                    bg_image_desktop_width = 160
                                    bg_image_desktop_height = 160
                                    bg_image_mobile_width = 192
                                    bg_image_mobile_height = 192

                                if card_number == '6 Card':
                                    bg_image_desktop_width = 160
                                    bg_image_desktop_height = 160
                                    bg_image_mobile_width = 192
                                    bg_image_mobile_height = 192

                                if (card_number == '7 Card (2-2-2-1)' and card_list == 'Right') or (
                                        card_number == '7 Card (2-2-1-2)' and card_list == 'Center Right') or (
                                        card_number == '7 Card (2-1-2-2)' and card_list == 'Center Left') or (
                                        card_number == '7 Card (1-2-2-2)' and card_list == 'Left'):
                                    bg_image_desktop_width = 282
                                    bg_image_desktop_height = 282
                                    bg_image_mobile_width = 264
                                    bg_image_mobile_height = 264
                                elif card_number == '7 Card (2-2-2-1)' or card_number == '7 Card (2-2-1-2)' or card_number == '7 Card (2-1-2-2)' or card_number == '7 Card (1-2-2-2)':
                                    bg_image_desktop_width = 160
                                    bg_image_desktop_height = 160
                                    bg_image_mobile_width = 192
                                    bg_image_mobile_height = 192

                                if card_number == '8 Card':
                                    bg_image_desktop_width = 160
                                    bg_image_desktop_height = 160
                                    bg_image_mobile_width = 192
                                    bg_image_mobile_height = 192
                            if " " in card_list:
                                card_size = "Small"
                            else:
                                card_size = "Large"
                            cell_value = process_label_text(exl_ws, card_compo_detail, col_location,
                                                            cell_area + " | " + card_list + "(" + card_size + ")",
                                                            "Text", "Desktop",
                                                            "N", "span",
                                                            "showcase-card-tab-card__product-name--desktop", "Y",
                                                            "Y", raw_data_meta)
                            cell_value = process_label_text(exl_ws, card_compo_detail, col_location,
                                                            cell_area + " | " + card_list + "(" + card_size + ")",
                                                            "Text", "Mobile",
                                                            "N", "span",
                                                            "showcase-card-tab-card__product-name--mobile", "Y",
                                                            "Y", raw_data_meta)

                            cell_value = process_label_text(exl_ws, card_compo_detail, col_location,
                                                            cell_area + " | " + card_list + "(" + card_size + ")",
                                                            "Text",
                                                            "Desktop",
                                                            "N", "span",
                                                            "showcase-card-tab-card__product-description--desktop",
                                                            "N",
                                                            "Y", raw_data_meta)
                            cell_value = process_label_text(exl_ws, card_compo_detail, col_location,
                                                            cell_area + " | " + card_list + "(" + card_size + ")",
                                                            "Text",
                                                            "Mobile",
                                                            "N", "span",
                                                            "showcase-card-tab-card__product-description--mobile",
                                                            "N",
                                                            "Y", raw_data_meta)
                            process_label_text_cta(exl_ws, card_compo_detail, col_location,
                                                   cell_area + " | " + card_list + "(" + card_size + ")", "CTA",
                                                   "N", "span.cta", raw_data_meta)

                            process_background_image(exl_ws, card_compo_detail, col_location,
                                                     cell_area + " | " + card_list + "(" + card_size + ")",
                                                     "BG Image",
                                                     "div.showcase-card-tab-card__img-wrap",
                                                     bg_image_desktop_width, bg_image_desktop_height,
                                                     bg_image_mobile_width, bg_image_mobile_height, "Y", "Y", raw_data_meta)

            if 'pd-g-feature-benefit-column-carousel' in content_comp_name:
                comp_tab_no = 0
                for comp_ft12_tab in content_comp_div.select(
                        'div > div.feature-column-carousel__column > div.feature-column-carousel__inner > div.feature-column-carousel__item'):
                    comp_tab_no += 1
                    cell_value = process_label_text(exl_ws, comp_ft12_tab, "", "Card " + str(comp_tab_no),
                                                    "Headline Text", "All",
                                                    "div > div > div > div.feature-column-carousel__title", "H",
                                                    "N", "Y", "Y", raw_data_meta)
                    cell_value = process_label_text(exl_ws, comp_ft12_tab, "", "Card " + str(comp_tab_no),
                                                    "Sub Headline Text", "All",
                                                    "div > div > div > div.feature-column-carousel__sub-title",
                                                    "H", "N", "Y", "Y", raw_data_meta)
                    cell_value = process_label_text(exl_ws, comp_ft12_tab, "", "Card " + str(comp_tab_no),
                                                    "Description",
                                                    "All",
                                                    "div > div > div > div.feature-column-carousel__text", "p",
                                                    "N", "N", "Y", raw_data_meta)
                    process_cta_buttons(exl_ws, comp_ft12_tab, "", "Card " + str(comp_tab_no), "CTA",
                                        "div > div > div.feature-column-carousel__button", raw_data_meta)
                    bg_image_desktop_width = 570
                    bg_image_desktop_height = 304
                    bg_image_mobile_width = 720
                    bg_image_mobile_height = 540
                    process_background_image(exl_ws, comp_ft12_tab, "", "Card " + str(comp_tab_no), "BG Image ",
                                             'div.feature-column-carousel__figure', bg_image_desktop_width,
                                             bg_image_desktop_height, bg_image_mobile_width, bg_image_mobile_height,
                                             "Y", "Y", raw_data_meta)

            if 'cm-g-bleed-card' in content_comp_name:
                comp_tab_no = 0
                # print(content_comp_div.select('div > div > div.swiper-wrapper'))
                for comp_co11_tab in content_comp_div.select(
                        'div > div > div.swiper-wrapper > div.bleed-card__item'):
                    comp_tab_no += 1
                if comp_tab_no == 2:
                    bg_image_desktop_width = 720
                    bg_image_desktop_height = 520
                else:
                    bg_image_desktop_width = 480
                    bg_image_desktop_height = 520
                if 'bleed-card--carousel-show-next' in content_comp_div.find('div').get('class', []):
                    bg_image_mobile_width = 624
                    bg_image_mobile_height = 850
                else:
                    bg_image_mobile_width = 720
                    bg_image_mobile_height = 850
                comp_tab_no = 0
                for comp_co11_tab in content_comp_div.select(
                        'div > div > div.swiper-wrapper > div.bleed-card__item'):
                    comp_tab_no += 1
                    cell_value = process_label_text(exl_ws, comp_co11_tab, "Bleed Card", "Card " + str(comp_tab_no),
                                                    "Headline Text", "All",
                                                    "div > div > div.bleed-card__text-wrap > div", "H",
                                                    "bleed-card__title", "Y", "Y", raw_data_meta)
                    cell_value = process_label_text(exl_ws, comp_co11_tab, "Bleed Card", "Card " + str(comp_tab_no),
                                                    "Sub Headline Text", "All",
                                                    "div > div > div.bleed-card__text-wrap > div", "H",
                                                    "bleed-card__sub-title", "N", "Y", raw_data_meta)
                    process_cta_buttons(exl_ws, comp_co11_tab, "Bleed Card", "Card " + str(comp_tab_no), "CTA",
                                        "div > div > div.bleed-card__text-wrap > div > div > div.cta-wrap", raw_data_meta)
                    process_background_image(exl_ws, comp_co11_tab, "Bleed Card", "Card " + str(comp_tab_no),
                                             "BG Image ",
                                             'div > div > div.bleed-card__image > div', bg_image_desktop_width,
                                             bg_image_desktop_height, bg_image_mobile_width, bg_image_mobile_height,
                                             "Y", "Y", raw_data_meta)

            if 'of-g-feature-benefit-card' in content_comp_name:
                comp_tab_no = 0
                # print(content_comp_div.select('div > div > div.swiper-wrapper'))
                for comp_co11_tab in content_comp_div.select(
                        'div > div > div.feature-benefit-card__card-wrap > div.feature-benefit-card__card'):
                    comp_tab_no += 1
                if 'feature-benefit-card--icon-image' in content_comp_div.select_one(
                        'div > div.feature-benefit-card__card-inner').get('class', []):
                    bg_image_desktop_width = 96
                    bg_image_desktop_height = 96
                    bg_image_mobile_width = 192
                    bg_image_mobile_height = 192
                else:
                    if comp_tab_no == 1:
                        bg_image_desktop_width = 720
                        bg_image_desktop_height = 520
                    elif comp_tab_no == 2:
                        bg_image_desktop_width = 684
                        bg_image_desktop_height = 440
                    elif comp_tab_no == 3:
                        bg_image_desktop_width = 448
                        bg_image_desktop_height = 440
                    else:
                        bg_image_desktop_width = 330
                        bg_image_desktop_height = 440
                    bg_image_mobile_width = 624
                    bg_image_mobile_height = 892
                comp_tab_no = 0
                for comp_co11_tab in content_comp_div.select(
                        'div > div > div.feature-benefit-card__card-wrap > div.feature-benefit-card__card'):
                    comp_tab_no += 1
                    cell_value = process_label_text(exl_ws, comp_co11_tab, "Feature Card",
                                                    "Card " + str(comp_tab_no),
                                                    "Eyeblow Text", "All",
                                                    "div > div.feature-benefit-card__text-wrap > div > div", "div",
                                                    "feature-benefit-card__eyebrow-text", "Y", "Y", raw_data_meta)
                    cell_value = process_label_text(exl_ws, comp_co11_tab, "Feature Card",
                                                    "Card " + str(comp_tab_no),
                                                    "Sub Headline Text", "All",
                                                    "div > div.feature-benefit-card__text-wrap > div > div", "H",
                                                    "feature-benefit-card__title", "Y", "Y", raw_data_meta)
                    cell_value = process_label_text(exl_ws, comp_co11_tab, "Feature Card",
                                                    "Card " + str(comp_tab_no),
                                                    "Description", "All",
                                                    "div > div.feature-benefit-card__text-wrap > div > div", "div",
                                                    "feature-benefit-card__description", "N", "Y", raw_data_meta)
                    process_cta_buttons(exl_ws, comp_co11_tab, "Feature Card", "Card " + str(comp_tab_no), "CTA",
                                        "div > div.feature-benefit-card__text-wrap > div > div > div.feature-benefit-card__cta", raw_data_meta)
                    process_background_image(exl_ws, comp_co11_tab, "Feature Card", "Card " + str(comp_tab_no),
                                             "BG Image ",
                                             'div > div.feature-benefit-card__figure > div', bg_image_desktop_width,
                                             bg_image_desktop_height, bg_image_mobile_width, bg_image_mobile_height,
                                             "Y", "Y", raw_data_meta)

            if 'pd-g-header-carousel' in content_comp_name:
                comp_tab_no = 0
                bg_image_desktop_width = 1440
                bg_image_desktop_height = 640
                bg_image_mobile_width = 720
                bg_image_mobile_height = 1064
                for comp_co11_tab in content_comp_div.select(
                        'section > div > div > div.header-carousel__slide'):
                    comp_tab_no += 1

                    cell_value = process_label_text(exl_ws, comp_co11_tab, "KeyVisual",
                                                    "Carousel " + str(comp_tab_no),
                                                    "Header Text", "All",
                                                    "section > div > div > div.header-carousel__slide > div > div > div.header-carousel__text-inner",
                                                    "H",
                                                    "header-carousel__headline", "Y", "Y", raw_data_meta)
                    cell_value = process_label_text(exl_ws, comp_co11_tab, "KeyVisual",
                                                    "Carousel " + str(comp_tab_no),
                                                    "Description", "All",
                                                    "section > div > div > div.header-carousel__slide > div > div > div.header-carousel__text-inner",
                                                    "p",
                                                    "header-carousel__description", "N", "Y", raw_data_meta)
                    process_cta_buttons(exl_ws, comp_co11_tab, "KeyVisual", "Carousel " + str(comp_tab_no), "CTA",
                                        "section > div > div > div.header-carousel__slide > div > div > div.header-carousel__text-inner > div.header-carousel__cta-wrap", raw_data_meta)
                    process_background_image(exl_ws, comp_co11_tab, "KeyVisual", "Carousel " + str(comp_tab_no),
                                             "BG Image ",
                                             'section > div > div > div.header-carousel__slide > div > div.header-carousel__background-media > div',
                                             bg_image_desktop_width,
                                             bg_image_desktop_height, bg_image_mobile_width, bg_image_mobile_height,
                                             "Y", "Y", raw_data_meta)

            if 'ho-g-showcase-card' in content_comp_name:

                col_location = 'Showcase Card'

                card_layout_option = ''
                card_no = 0

                for card_compo in content_comp_div.find_all("div", class_='co35-showcase-card-tab-card'):
                    # print(card_compo)
                    if 'co35-showcase-card-tab-card--large' in card_compo.attrs[
                        'class'] or 'co35-showcase-card-tab-card--product-large' in card_compo.attrs[
                        'class'] or 'co35-showcase-card-tab-card--vertical' in card_compo.attrs[
                        'class'] or 'co35-showcase-card-tab-card--product-vertical' in card_compo.attrs['class']:
                        card_layout_option += 'L'
                    elif 'co35-showcase-card-tab-card--small' in card_compo.attrs[
                        'class'] or 'co35-showcase-card-tab-card--product-small' in card_compo.attrs['class']:
                        card_layout_option += 'S'

                options = {'SSSSSSSS': ['8 Card',
                                        ['Top Left', 'Bottom Left', 'Top Center Left', 'Bottom Center Left',
                                         'Top Center Right', 'Bottom Center Right', 'Top Right',
                                         'Bottom Right']],
                           'LSSSSSS': ['7 Card (1-2-2-2)',
                                       ['Left', 'Top Center Left', 'Bottom Center Left', 'Top Center Right',
                                        'Bottom Center Right', 'Top Right', 'Bottom Right']],
                           'SSLSSSS': ['7 Card (2-1-2-2)',
                                       ['Top Left', 'Bottom Left', 'Center Left', 'Top Center Right',
                                        'Bottom Center Right', 'Top Right', 'Bottom Right']],
                           'SSSSLSS': ['7 Card (2-2-1-2)',
                                       ['Top Left', 'Bottom Left', 'Top Center Left', 'Bottom Center Left',
                                        'Center Right', 'Top Right', 'Bottom Right']],
                           'SSSSSSL': ['7 Card (2-2-2-1)',
                                       ['Top Left', 'Bottom Left', 'Top Center Left', 'Bottom Center Left',
                                        'Top Center Right', 'Bottom Center Right', 'Right']],
                           'SSSSSS': ['6 Card',
                                      ['Top Left', 'Bottom Left', 'Top Center', 'Bottom Center', 'Top Right',
                                       'Bottom Right']],
                           'SSSSL': ['5 Card (2-2-1)',
                                     ['Top Left', 'Bottom Left', 'Top Center', 'Bottom Center', 'Right']],
                           'SSLSS': ['5 Card (2-1-2)',
                                     ['Top Left', 'Bottom Left', 'Center', 'Top Right', 'Bottom Right']],
                           'LSSSS': ['5 Card (1-2-2)',
                                     ['Left', 'Top Center', 'Bottom Center', 'Top Right', 'Bottom Right']],
                           'LLL': ['3 Card', ['Left', 'Center', 'Right']]}
                for x in options.keys():
                    if x == card_layout_option:
                        card_number = options[x][0]
                        card_layout = options[x][1]
                card_list_no = 0
                for card_list in card_layout:
                    card_list_no += 1

                    card_compo_detail = content_comp_div.find_all("div", class_='co35-showcase-card-tab-card')[
                        card_list_no - 1]
                    card_type = process_class_attributes(exl_ws, card_compo_detail, "", "", "", "",
                                                         'N', 'N',
                                                         {"co35-showcase-card-tab-card--large": "Full Breed",
                                                          "co35-showcase-card-tab-card--small": "Full Breed",
                                                          "co35-showcase-card-tab-card--vertical": "Full Breed",
                                                          "co35-showcase-card-tab-card--product-large": "Product",
                                                          "co35-showcase-card-tab-card--product-small": "Product",
                                                          "co35-showcase-card-tab-card--product-vertical": "Product"},
                                                         'N',
                                                         "N", raw_data_meta)

                    if card_type == 'Full Breed':
                        if card_number == '3 Card':
                            bg_image_desktop_width = 448
                            bg_image_desktop_height = 684
                            bg_image_mobile_width = 624
                            bg_image_mobile_height = 352

                        if (card_number == '5 Card (1-2-2)' and card_list == 'Left') or (
                                card_number == '5 Card (2-1-2)' and card_list == 'Center') or (
                                card_number == '5 Card (2-2-1)' and card_list == 'Right'):
                            bg_image_desktop_width = 684
                            bg_image_desktop_height = 684
                            bg_image_mobile_width = 624
                            bg_image_mobile_height = 352
                        elif card_number == '5 Card (1-2-2)' or card_number == '5 Card (2-1-2)' or card_number == '5 Card (2-2-1)':
                            bg_image_desktop_width = 330
                            bg_image_desktop_height = 330
                            bg_image_mobile_width = 296
                            bg_image_mobile_height = 352

                        if card_number == '6 Card':
                            bg_image_desktop_width = 448
                            bg_image_desktop_height = 330
                            bg_image_mobile_width = 296
                            bg_image_mobile_height = 352

                        if (card_number == '7 Card (2-2-2-1)' and card_list == 'Right') or (
                                card_number == '7 Card (2-2-1-2)' and card_list == 'Center Right') or (
                                card_number == '7 Card (2-1-2-2)' and card_list == 'Center Left') or (
                                card_number == '7 Card (1-2-2-2)' and card_list == 'Left'):
                            bg_image_desktop_width = 330
                            bg_image_desktop_height = 684
                            bg_image_mobile_width = 624
                            bg_image_mobile_height = 352
                        elif card_number == '7 Card (2-2-2-1)' or card_number == '7 Card (2-2-1-2)' or card_number == '7 Card (2-1-2-2)' or card_number == '7 Card (1-2-2-2)':
                            bg_image_desktop_width = 330
                            bg_image_desktop_height = 330
                            bg_image_mobile_width = 296
                            bg_image_mobile_height = 352

                        if card_number == '8 Card':
                            bg_image_desktop_width = 330
                            bg_image_desktop_height = 330
                            bg_image_mobile_width = 296
                            bg_image_mobile_height = 352

                    if card_type == 'Product':
                        if card_number == '3 Card':
                            bg_image_desktop_width = 376
                            bg_image_desktop_height = 376
                            bg_image_mobile_width = 264
                            bg_image_mobile_height = 264

                        if (card_number == '5 Card (1-2-2)' and card_list == 'Left') or (
                                card_number == '5 Card (2-1-2)' and card_list == 'Center') or (
                                card_number == '5 Card (2-2-1)' and card_list == 'Right'):
                            bg_image_desktop_width = 376
                            bg_image_desktop_height = 376
                            bg_image_mobile_width = 264
                            bg_image_mobile_height = 264
                        elif card_number == '5 Card (1-2-2)' or card_number == '5 Card (2-1-2)' or card_number == '5 Card (2-2-1)':
                            bg_image_desktop_width = 160
                            bg_image_desktop_height = 160
                            bg_image_mobile_width = 192
                            bg_image_mobile_height = 192

                        if card_number == '6 Card':
                            bg_image_desktop_width = 160
                            bg_image_desktop_height = 160
                            bg_image_mobile_width = 192
                            bg_image_mobile_height = 192

                        if (card_number == '7 Card (2-2-2-1)' and card_list == 'Right') or (
                                card_number == '7 Card (2-2-1-2)' and card_list == 'Center Right') or (
                                card_number == '7 Card (2-1-2-2)' and card_list == 'Center Left') or (
                                card_number == '7 Card (1-2-2-2)' and card_list == 'Left'):
                            bg_image_desktop_width = 282
                            bg_image_desktop_height = 282
                            bg_image_mobile_width = 264
                            bg_image_mobile_height = 264
                        elif card_number == '7 Card (2-2-2-1)' or card_number == '7 Card (2-2-1-2)' or card_number == '7 Card (2-1-2-2)' or card_number == '7 Card (1-2-2-2)':
                            bg_image_desktop_width = 160
                            bg_image_desktop_height = 160
                            bg_image_mobile_width = 192
                            bg_image_mobile_height = 192

                        if card_number == '8 Card':
                            bg_image_desktop_width = 160
                            bg_image_desktop_height = 160
                            bg_image_mobile_width = 192
                            bg_image_mobile_height = 192
                    if " " in card_list:
                        card_size = "Small"
                    else:
                        card_size = "Large"
                    cell_value = process_label_text(exl_ws, card_compo_detail, col_location,
                                                    card_list + "(" + card_size + ")",
                                                    "Text", "Desktop",
                                                    "N", "span",
                                                    "co35-showcase-card-tab-card__product-name--desktop", "Y", "Y", raw_data_meta)
                    cell_value = process_label_text(exl_ws, card_compo_detail, col_location,
                                                    card_list + "(" + card_size + ")",
                                                    "Text", "Mobile",
                                                    "N", "span",
                                                    "co35-showcase-card-tab-card__product-name--mobile", "Y", "Y", raw_data_meta)

                    cell_value = process_label_text(exl_ws, card_compo_detail, col_location,
                                                    card_list + "(" + card_size + ")",
                                                    "Text",
                                                    "Desktop",
                                                    "N", "span",
                                                    "co35-showcase-card-tab-card__product-description--desktop",
                                                    "N",
                                                    "Y", raw_data_meta)
                    cell_value = process_label_text(exl_ws, card_compo_detail, col_location,
                                                    card_list + "(" + card_size + ")",
                                                    "Text",
                                                    "Mobile",
                                                    "N", "span",
                                                    "co35-showcase-card-tab-card__product-description--mobile", "N",
                                                    "Y", raw_data_meta)
                    process_label_text_cta(exl_ws, card_compo_detail, col_location,
                                           card_list + "(" + card_size + ")", "CTA",
                                           "N", "span.cta", raw_data_meta)

                    process_background_image(exl_ws, card_compo_detail, col_location,
                                             card_list + "(" + card_size + ")",
                                             "BG Image",
                                             "div.co35-showcase-card-tab-card__img-wrap",
                                             bg_image_desktop_width, bg_image_desktop_height,
                                             bg_image_mobile_width, bg_image_mobile_height, "Y", "Y", raw_data_meta)
            if 'pd-g-feature-benefit-banner' in content_comp_name:
                if 'st-feature-benefit-banner--full' in content_comp_div.select_one('section').get('class', []):
                    bg_image_desktop_width = 1440
                    bg_image_desktop_height = 304
                    bg_image_mobile_width = 720
                    bg_image_mobile_height = 540
                    col_area = "Full Image Banner"
                else:
                    # print("half")
                    bg_image_desktop_width = 570
                    bg_image_desktop_height = 304
                    bg_image_mobile_width = 570
                    bg_image_mobile_height = 304
                    col_area = "Half Image Banner"

                cell_value = process_label_text(exl_ws, content_comp_div, "Banner", col_area, "Headline Text",
                                                "Desktop",
                                                "section > div > div > div.st-feature-benefit-banner__content-wrap",
                                                "H", "desktop-only", "Y", "Y", raw_data_meta)
                cell_value = process_label_text(exl_ws, content_comp_div, "Banner", col_area, "Headline Text",
                                                "Mobile",
                                                "section > div > div > div.st-feature-benefit-banner__content-wrap",
                                                "H", "mobile-only", "Y", "Y", raw_data_meta)
                cell_value = process_label_text(exl_ws, content_comp_div, "Banner", col_area, "Description Text",
                                                "All",
                                                "section > div > div > div.st-feature-benefit-banner__desc",
                                                "p", "desktop-only", "N", "Y", raw_data_meta)
                process_label_text_cta(exl_ws, content_comp_div, "Banner", col_area, "CTA",
                                       "section > div > div > div.st-feature-benefit-banner__content-wrap > div.st-feature-benefit-banner__cta-wrap",
                                       "a.cta", raw_data_meta)
                process_background_image(exl_ws, content_comp_div, "Banner", col_area, "BG Image ",
                                         'section > div > figure.st-feature-benefit-banner__figure > div',
                                         bg_image_desktop_width,
                                         bg_image_desktop_height, bg_image_mobile_width, bg_image_mobile_height,
                                         "Y", "Y", raw_data_meta)

            if 'co-g-showcase-card-tab' in content_comp_name:
                col_location = process_label_text(exl_ws, content_comp_div, "", " ",
                                                  "", "",
                                                  "div.co-showcase-card-tab",
                                                  "H",
                                                  "co-showcase-card-tab__headline", "N", "N", raw_data_meta)
                co19_area_name = content_comp_div.select('div > div.co-showcase-card-tab__tabs > div > button')
                co19_content_div = content_comp_div.select(
                    'div > div.co-showcase-card-tab__wrap > div > section.co-showcase-card-tab__container > div')
                no_co19_tab = 0
                for co19_col_div in co19_content_div:
                    no_co19_tab += 1
                    col_area = co19_area_name[no_co19_tab - 1].text.strip()

                    card_layout_option = ''
                    card_no = 0
                    for card_compo in co19_col_div.select('div.co-showcase-card-tab__card'):
                        if 'co-showcase-card-tab__card--hero' in card_compo.attrs['class']:
                            card_layout_option += 'L'
                        elif 'co-showcase-card-tab__card--small' in card_compo.attrs['class']:
                            card_layout_option += 'S'
                    # print(card_layout_option)
                    options = {'LSS': ['3 Card',
                                       ['Left', 'Top Right', 'Bottom Right']],
                               'SSL': ['3 Card',
                                       ['Top Left', 'Bottom Left', 'Right']]}
                    for x in options.keys():
                        if x == card_layout_option:
                            card_number = options[x][0]
                            card_layout = options[x][1]

                    for card_list in card_layout:
                        card_no += 1
                        if card_list == "Left" or card_list == "Right":
                            cell_value1 = process_label_text(exl_ws,
                                                             co19_col_div.select('div.co-showcase-card-tab__card')[
                                                                 card_no - 1], col_location,
                                                             col_area + " | " + card_list, "Headline Text", "All",
                                                             "div.co-showcase-card-tab__text", "p",
                                                             "co-showcase-card-tab__card-headline", "Y", "Y", raw_data_meta)
                            cell_value2 = process_label_text(exl_ws,
                                                             co19_col_div.select('div.co-showcase-card-tab__card')[
                                                                 card_no - 1], col_location,
                                                             col_area + " | " + card_list, "Sub Headline Text",
                                                             "All",
                                                             "div.co-showcase-card-tab__text", "p",
                                                             "co-showcase-card-tab__card-sub-headline", "N", "Y", raw_data_meta)
                            bg_image_desktop_width = 802
                            bg_image_desktop_height = 520
                            bg_image_mobile_width = 624
                            bg_image_mobile_height = 624
                        else:
                            cell_value1 = process_label_text(exl_ws,
                                                             co19_col_div.select('div.co-showcase-card-tab__card')[
                                                                 card_no - 1], col_location,
                                                             col_area + " | " + card_list, "Headline Text", "All",
                                                             "N", "p",
                                                             "co-showcase-card-tab__card-headline", "Y", "Y", raw_data_meta)
                            cell_value2 = process_label_text(exl_ws,
                                                             co19_col_div.select('div.co-showcase-card-tab__card')[
                                                                 card_no - 1], col_location,
                                                             col_area + " | " + card_list, "Sub Headline Text",
                                                             "All",
                                                             "N", "p",
                                                             "co-showcase-card-tab__card-sub-headline", "N", "Y", raw_data_meta)
                            bg_image_desktop_width = 128
                            bg_image_desktop_height = 128
                            bg_image_mobile_width = 144
                            bg_image_mobile_height = 144

                        process_background_image(exl_ws,
                                                 co19_col_div.select('div.co-showcase-card-tab__card')[card_no - 1],
                                                 col_location,
                                                 col_area + " | " + card_list, "BG Image", 'div.figure',
                                                 bg_image_desktop_width, bg_image_desktop_height,
                                                 bg_image_mobile_width, bg_image_mobile_height, "Y", "Y", raw_data_meta)

            if 'pd-g-feature-benefit-two-column' in content_comp_name:
                ft09_components = content_comp_div.select('section > div > div.st-two-column__column')
                ft09_com_no = 0
                for ft09_component in ft09_components:
                    ft09_com_no += 1
                    cell_value = process_label_text(exl_ws, ft09_component, "2 Column",
                                                    'Column ' + str(ft09_com_no), "Headline Text", "All",
                                                    "div.st-two-column__title", "H",
                                                    "st-two-column__headline", "Y", "Y", raw_data_meta)
                    cell_value = process_label_text(exl_ws, ft09_component, "2 Column",
                                                    'Column ' + str(ft09_com_no), "Sub Title Text", "All",
                                                    "N", "div",
                                                    "st-two-column__sub-title", "N", "Y", raw_data_meta)
                    cell_value = process_label_text(exl_ws, ft09_component, "2 Column",
                                                    'Column ' + str(ft09_com_no), "Description", "All",
                                                    "div.st-two-column__text", "p",
                                                    "N", "N", "Y", raw_data_meta)

                    bg_image_desktop_width = 590
                    bg_image_desktop_height = 0
                    bg_image_mobile_width = 624
                    bg_image_mobile_height = 0

                    process_label_text_cta(exl_ws, ft09_component, "2 Column", 'Column ' + str(ft09_com_no), "CTA",
                                           "div.st-two-column__cta-wrap", "a.cta", raw_data_meta)
                    process_background_image(exl_ws, ft09_component, "2 Column", 'Column ' + str(ft09_com_no),
                                             "BG Image ",
                                             'figure.st-two-column__figure', bg_image_desktop_width,
                                             bg_image_desktop_height, bg_image_mobile_width, bg_image_mobile_height,
                                             "Y", "Y", raw_data_meta)

            if 'pd-g-feature-benefit-full-bleed' in content_comp_name:
                ft03_components = content_comp_div.select('section > div.st-feature-benefit-full-bleed__wrap')
                ft03_compo_width = content_comp_div.find('section').get('class', [])
                if 'st-feature-benefit-full-bleed--width-1920' in ft03_compo_width:
                    bg_image_desktop_width = 1920
                    bg_image_desktop_height = 0
                else:
                    bg_image_desktop_width = 1440
                    bg_image_desktop_height = 0

                if 'st-feature-benefit-full-bleed--image-radius' in ft03_compo_width:
                    bg_image_mobile_width = 624
                    bg_image_mobile_height = 0
                else:
                    bg_image_mobile_width = 720
                    bg_image_mobile_height = 0
                for ft03_component in ft03_components:
                    cell_value1 = process_label_text(exl_ws, ft03_component, "Banner",
                                                     "Column 1", "Headline Text", "All",
                                                     "div > div > div.st-feature-benefit-full-bleed__content-area",
                                                     "H",
                                                     "st-feature-benefit-full-bleed__title", "Y", "Y", raw_data_meta)
                    cell_value2 = process_label_text(exl_ws, ft03_component, "Banner",
                                                     "Column 1", "Sub Headline Text", "All",
                                                     "div > div > div.st-feature-benefit-full-bleed__content-area",
                                                     "H",
                                                     "st-feature-benefit-full-bleed__sub-title", "N", "Y", raw_data_meta)
                    cell_value3 = process_label_text(exl_ws, ft03_component, "Banner",
                                                     "Column 1", "Description", "All",
                                                     "div > div > div.st-feature-benefit-full-bleed__content-area",
                                                     "p",
                                                     "st-feature-benefit-full-bleed__text", "N", "Y", raw_data_meta)
                    process_label_text_cta(exl_ws, ft03_component, "Banner", "Column 1", "CTA",
                                           "div > div > div > div.st-feature-benefit-full-bleed__cta", "a.cta", raw_data_meta)

                    process_background_image(exl_ws, ft03_component, "Banner", "Column 1", "BG Image ",
                                             'figure.st-feature-benefit-full-bleed__figure ',
                                             bg_image_desktop_width,
                                             bg_image_desktop_height, bg_image_mobile_width, bg_image_mobile_height,
                                             "Y",
                                             "Y", raw_data_meta)

            if 'nv-g-local-floating-nav' in content_comp_name:
                bg_image_desktop_width = 140
                bg_image_desktop_height = 140
                bg_image_mobile_width = 140
                bg_image_mobile_height = 140
                nv06_components = content_comp_div.select('div > div > nav > ul > li')
                nv06_no = 0
                for nv06_component in nv06_components:
                    nv06_no += 1
                    cell_value = process_label_text(exl_ws, nv06_component, "Category Browser",
                                                    "Category " + str(nv06_no), "Headline Text", "All",
                                                    "a", "div",
                                                    "nv-local-floating-nav__headline", "Y", "Y", raw_data_meta)
                    process_background_image(exl_ws, nv06_component, "Category Browser", "Category " + str(nv06_no),
                                             "BG Image ",
                                             'a.nv-local-floating-nav__content', bg_image_desktop_width,
                                             bg_image_desktop_height, bg_image_mobile_width, bg_image_mobile_height,
                                             "Y",
                                             "Y", raw_data_meta)

        row_num = exl_ws.max_row + 1
        for i in range(1, 9):
            exl_ws.cell(row=row_num, column=i).border = Border(top=Side(style='thin', color='555555'))

        df = pd.DataFrame(data_list)
        file_name_base = url.replace('https://www.samsung.com/', '').replace('/', '-')
        file_name_base = file_name_base.strip('-')
        current_time = datetime.now().strftime("%Y%m%d%H%M%S")
        end_time = datetime.now()

        path = "./result/" + end_time.strftime("%Y%m%d")
        os.makedirs(path, exist_ok=True)
        final_file_name = f"{path}/{file_name_base}_{current_time}.xlsx"
        exl_wb.save(final_file_name)

        time_elapsed = end_time - start_time
        print("FINISH", url, 'at : ', end_time.strftime("%Y/%m/%d %H:%M:%S"), "(", time_elapsed, ")")
        print(f"SAVE AT : {final_file_name}")


    except Exception as e:
        error_logger(e,url,start_time)
        # file_name_base = url.replace('https://www.samsung.com/', '').replace('/', '-')
        # file_name_base = file_name_base.strip('-')
        # today = datetime.now().strftime("%Y%m%d")
        # path = "./result/" + today + "/error"
        # os.makedirs(path, exist_ok=True)
        # current_time = datetime.now().strftime("%Y%m%d%H%M%S")
        # final_file_name = f"{path}/{file_name_base}_eroror_{current_time}.txt"
        #
        # end_time = datetime.now()
        # time_elapsed = end_time - start_time
        #
        # f = open(final_file_name, "w")
        #
        # error_log = f"{e}, \n ERROR :, {url}, at : {end_time} ({time_elapsed})"
        # f.write(error_log)
        # f.close()
        #
        # print("ERROR :", url, 'at : ', end_time.strftime("%Y/%m/%d %H:%M:%S"), "(", time_elapsed, ")")
        # print(error_log)