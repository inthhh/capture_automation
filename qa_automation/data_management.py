import pandas as pd
import ast
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Alignment, Border, Side
# from db import execute
from datetime import datetime
import os

data_list = []
before_location = ''
before_area = ''
before_title = ''


def add_data_to_list(exl_ws, col_location, col_area, col_title, cell_description, cell_contents, cell_check, cell_inspection, raw_data_meta, db_conn):


    global data_list, before_location, before_area, before_title

    tab_menu_changed = col_location != before_location

    cell_location = '' if col_location == before_location else col_location
    cell_area = col_area if tab_menu_changed or col_area != before_area else ''
    cell_title = '' if col_title == before_title else col_title

    saved_item = [cell_location, cell_area, cell_title, cell_description, cell_contents, cell_check, cell_inspection]
    # print(col_location, col_area, col_title, cell_description, cell_check, cell_inspection)
    exl_ws.append(saved_item)

    row_num = exl_ws.max_row

    if cell_location:
        exl_ws.cell(row=row_num, column=1).border = Border(top=Side(style='thin', color='555555'), right=Side(style='thin', color='555555'))
    else:
        exl_ws.cell(row=row_num, column=1).border = Border(right=Side(style='thin', color='555555'))
    exl_ws.cell(row=row_num, column=1).alignment = Alignment(vertical='center')

    if cell_area:
        exl_ws.cell(row=row_num, column=2).border = Border(top=Side(style='thin', color='555555'), right=Side(style='thin', color='555555'))
    else:
        exl_ws.cell(row=row_num, column=2).border = Border(right=Side(style='thin', color='555555'))
    exl_ws.cell(row=row_num, column=2).alignment = Alignment(vertical='center')

    if cell_title:
        exl_ws.cell(row=row_num, column=3).border = Border(top=Side(style='thin', color='555555'), right=Side(style='thin', color='555555'))
    else:
        exl_ws.cell(row=row_num, column=3).border = Border(right=Side(style='thin', color='555555'))
    exl_ws.cell(row=row_num, column=3).alignment = Alignment(vertical='center')

    exl_ws.cell(row=row_num, column=4).border = Border(top=Side(style='thin', color='555555'), right=Side(style='thin', color='555555'))
    exl_ws.cell(row=row_num, column=4).alignment = Alignment(vertical='center')
    exl_ws.cell(row=row_num, column=5).border = Border(top=Side(style='thin', color='555555'), right=Side(style='thin', color='555555'))
    exl_ws.cell(row=row_num, column=5).alignment = Alignment(vertical='center', wrap_text=True)
    exl_ws.cell(row=row_num, column=6).border = Border(top=Side(style='thin', color='555555'), right=Side(style='thin', color='555555'))
    exl_ws.cell(row=row_num, column=6).alignment = Alignment(vertical='center', horizontal='center')
    exl_ws.cell(row=row_num, column=7).border = Border(top=Side(style='thin', color='555555'), right=Side(style='thin', color='555555'))
    exl_ws.cell(row=row_num, column=7).alignment = Alignment(vertical='center', wrap_text=True)
    exl_ws.cell(row=row_num, column=8).border = Border(top=Side(style='thin', color='555555'), right=Side(style='thin', color='555555'))


    before_location = col_location
    before_area = col_area
    before_title = col_title
    insert_query = "INSERT INTO qa_result (date,rhq, subsidiary, site_code, page_type, category, location, area, title, description, contents, check_result, check_reason) values (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);"
    db_conn.execute(insert_query, (
        raw_data_meta["date"], raw_data_meta["rhq"], raw_data_meta["subsidiary"], raw_data_meta["site_code"],
        raw_data_meta["page_type"],raw_data_meta["category"], col_location, col_area, col_title, cell_description, cell_contents, cell_check,
        cell_inspection))
    date_now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(date_now,": ", col_location, col_area, col_title, cell_contents, cell_check, cell_inspection)


def add_dataimage_to_list(exl_ws, col_location, col_area, col_title, cell_description, cell_contents, cell_check, cell_inspection, image_bytes_resized, new_height, raw_data_meta, db_conn):
    # print(col_location, col_area, col_title, cell_description, cell_contents, cell_check, cell_inspection)

    global data_list, before_location, before_area, before_title

    tab_menu_changed = col_location != before_location

    cell_location = '' if col_location == before_location else col_location
    cell_area = col_area if tab_menu_changed or col_area != before_area else ''
    cell_title = '' if col_title == before_title else col_title

    saved_item = [cell_location, cell_area, cell_title, cell_description, cell_contents, cell_check, cell_inspection]
    exl_ws.append(saved_item)
    # print(col_location, col_area, col_title, cell_description, cell_check, cell_inspection)
    row_num = exl_ws.max_row

    if cell_location:
        exl_ws.cell(row=row_num, column=1).border = Border(top=Side(style='thin', color='555555'),
                                                           right=Side(style='thin', color='555555'))
    else:
        exl_ws.cell(row=row_num, column=1).border = Border(right=Side(style='thin', color='555555'))
    exl_ws.cell(row=row_num, column=1).alignment = Alignment(vertical='center')

    if cell_area:
        exl_ws.cell(row=row_num, column=2).border = Border(top=Side(style='thin', color='555555'),
                                                           right=Side(style='thin', color='555555'))
    else:
        exl_ws.cell(row=row_num, column=2).border = Border(right=Side(style='thin', color='555555'))
    exl_ws.cell(row=row_num, column=2).alignment = Alignment(vertical='center')

    if cell_title:
        exl_ws.cell(row=row_num, column=3).border = Border(top=Side(style='thin', color='555555'),
                                                           right=Side(style='thin', color='555555'))
    else:
        exl_ws.cell(row=row_num, column=3).border = Border(right=Side(style='thin', color='555555'))
    exl_ws.cell(row=row_num, column=3).alignment = Alignment(vertical='center')

    exl_ws.cell(row=row_num, column=4).border = Border(top=Side(style='thin', color='555555'),
                                                       right=Side(style='thin', color='555555'))
    exl_ws.cell(row=row_num, column=4).alignment = Alignment(vertical='center')
    exl_ws.cell(row=row_num, column=5).border = Border(top=Side(style='thin', color='555555'),
                                                       right=Side(style='thin', color='555555'))
    exl_ws.cell(row=row_num, column=5).alignment = Alignment(vertical='center', wrap_text=True)
    exl_ws.cell(row=row_num, column=6).border = Border(top=Side(style='thin', color='555555'),
                                                       right=Side(style='thin', color='555555'))
    exl_ws.cell(row=row_num, column=6).alignment = Alignment(vertical='center', horizontal='center')
    exl_ws.cell(row=row_num, column=7).border = Border(top=Side(style='thin', color='555555'),
                                                       right=Side(style='thin', color='555555'))
    exl_ws.cell(row=row_num, column=7).alignment = Alignment(vertical='center')
    exl_ws.cell(row=row_num, column=8).border = Border(top=Side(style='thin', color='555555'),
                                                       right=Side(style='thin', color='555555'))
    exl_ws.cell(row=row_num, column=8).alignment = Alignment(vertical='center')

    if image_bytes_resized:
        img = Image(image_bytes_resized)
        img.anchor = f'H{row_num}'
        exl_ws.row_dimensions[row_num].height = new_height * 0.75 + 1
        exl_ws.add_image(img, f'H{row_num}')


    before_location = col_location
    before_area = col_area
    before_title = col_title



    insert_query = "INSERT INTO qa_result (date,rhq, subsidiary, site_code, page_type, category, location, area,title, description, contents, check_result, check_reason) values (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);"
    db_conn.execute(insert_query, (
        raw_data_meta["date"], raw_data_meta["rhq"], raw_data_meta["subsidiary"], raw_data_meta["site_code"],
        raw_data_meta["page_type"], raw_data_meta["category"], col_location, col_area, col_title, cell_description, cell_contents,
        cell_check,
        cell_inspection))
    date_now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(date_now,": ", before_location, col_location, col_area, col_title, cell_contents, cell_check, cell_inspection)


# Error 발생시 로그 생성 메소드
def error_logger (error_message, url, start_time):
    file_name_base = url.replace('https://www.samsung.com/', '').replace('/', '-')
    file_name_base = file_name_base.strip('-')
    today = datetime.now().strftime("%Y%m%d")
    path = "./result/" + today + "/error"
    os.makedirs(path, exist_ok=True)
    current_time = datetime.now().strftime("%Y%m%d%H%M%S")
    final_file_name = f"{path}/{file_name_base}_error_{current_time}.txt"

    end_time = datetime.now()
    time_elapsed = end_time - start_time

    f = open(final_file_name, "w")

    error_log = f"{error_message}, \n ERROR :, {url}, at : {end_time} ({time_elapsed})"
    f.write(error_log)
    f.close()

    print("ERROR :", url, 'at : ', end_time.strftime("%Y/%m/%d %H:%M:%S"), "(", time_elapsed, ")")
    print(error_log)