import requests
import csv

from bs4 import BeautifulSoup
from datetime import datetime
# from pandas import ExcelWriter
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
import pandas as pd
from process_components import *
from data_management import *
from region import region

region = region

def qa (url_code):
    raw_data_meta = {
        "date": datetime.today().strftime('%Y-%m-%d'),
        "rhq": region[url_code]["rhq"],
        "subsidiary": region[url_code]["subsidiary"],
        "site_code": url_code,
        "page_type": "Home",
        "category": "Home",
    }

    url = 'https://www.samsung.com/' + url_code + '/'
    sitecode = url.split('/')[3]
    response = requests.get(url)
    html_all = BeautifulSoup(response.text, 'lxml')
    start_time = datetime.now()
    print("START", url, 'at : ', start_time.strftime("%Y/%m/%d %H:%M:%S"))
    try:
        content_all = html_all.select(
            '#content > div > div > div.responsivegrid.aem-GridColumn.aem-GridColumn--default--12 > div > div')

        exl_wb = Workbook()
        exl_ws = exl_wb.active
        exl_ws.title = "Home"

        exl_header = ['Location', 'Area', 'Title', 'Description', 'Contents', 'Check', 'Guide','Image']
        exl_ws.append(exl_header)

        exl_ws.column_dimensions['A'].width = 15
        exl_ws.column_dimensions['B'].width = 40
        exl_ws.column_dimensions['C'].width = 15
        exl_ws.column_dimensions['D'].width = 15
        exl_ws.column_dimensions['E'].width = 50
        exl_ws.column_dimensions['F'].width = 5
        exl_ws.column_dimensions['G'].width = 15
        exl_ws.column_dimensions['H'].width = 26
        exl_ws.sheet_view.showGridLines = False
        exl_ws.freeze_panes = 'A2'

        no_content_comp_div = 0
        for content_comp_div in content_all:
            no_content_comp_div += 1
            content_comp_name = content_comp_div.get('class', [])
            if 'ho-g-home-kv-carousel' in content_comp_name:

                cell_value = process_class_attributes(exl_ws, content_comp_div,
                                                      "", "", "", "",
                                                      "N", "section",
                                                      {"home-kv-carousel--width-large" : "1920px"}, "1440px","N",raw_data_meta)

                if cell_value == '1920px':
                    bg_image_desktop_width = 1920
                    bg_image_mobile_width = 720
                else:
                    bg_image_desktop_width = 1440
                    bg_image_mobile_width = 720

                cell_value = process_class_attributes(exl_ws, content_comp_div,
                                                      "", "", "", "",
                                                      "N", "section",
                                                      {"home-kv-carousel--height-large": "Desktop:810px / Mobile:640px",
                                                       "home-kv-carousel--height-medium": "Desktop:640px / Mobile:540px",
                                                       "home-kv-carousel--height-smedium": "Desktop:344px / Mobile:400px",
                                                       "home-kv-carousel--height-small": "Desktop:320px / Mobile:320px"}, "N", "N",raw_data_meta)
                if cell_value == 'Desktop:810px / Mobile:640px':
                    bg_image_desktop_height = 810
                    bg_image_mobile_height = 1280
                elif cell_value == 'Desktop:640px / Mobile:540px':
                    bg_image_desktop_height = 640
                    bg_image_mobile_height = 1080
                elif cell_value == 'Desktop:344px / Mobile:400px':
                    bg_image_desktop_height = 344
                    bg_image_mobile_height = 800
                elif cell_value == 'Desktop:320px / Mobile:320px':
                    bg_image_desktop_height = 320
                    bg_image_mobile_height = 640

                carousel_no = 0
                for content_carousel_div in content_comp_div.select(
                    'section > div.home-kv-carousel__container > div.home-kv-carousel__wrapper > div'):
                    carousel_no += 1

                    cell_value = process_attributes_value(exl_ws, content_carousel_div,
                                                         "KeyVisual", "KV" + str(carousel_no),"Headline Text", 'Desktop',
                                                         "div > div > div.home-kv-carousel__text-wrap", "H",
                                                         "home-kv-carousel__headline",
                                                         "data-desktop-headline-text", "Y", "Y",raw_data_meta)
                    cell_value = process_attributes_value(exl_ws, content_carousel_div,
                                                          "KeyVisual", "KV" + str(carousel_no), "Headline Text", 'Mobile',
                                                          "div > div > div.home-kv-carousel__text-wrap", "H",
                                                          "home-kv-carousel__headline",
                                                          "data-mobile-headline-text", "Y", "Y",raw_data_meta)
                    cell_value = process_attributes_value(exl_ws, content_carousel_div,
                                                          "KeyVisual", "KV" + str(carousel_no), "Description Text", 'Desktop',
                                                          "div > div > div.home-kv-carousel__text-wrap", "p",
                                                          "home-kv-carousel__desc",
                                                          "data-desktop-description", "N", "Y",raw_data_meta)
                    cell_value = process_attributes_value(exl_ws, content_carousel_div,
                                                          "KeyVisual", "KV" + str(carousel_no), "Description Text", 'Mobile',
                                                          "div > div > div.home-kv-carousel__text-wrap", "p",
                                                          "home-kv-carousel__desc",
                                                          "data-mobile-description", "N", "Y",raw_data_meta)

                    process_cta_buttons(exl_ws, content_carousel_div, "KeyVisual", "KV" + str(carousel_no), "CTA",
                                        "div > div > div.home-kv-carousel__text-wrap > div.home-kv-carousel__cta-wrap",raw_data_meta)

                    process_background_image(exl_ws, content_carousel_div, "KeyVisual", "KV" + str(carousel_no), "BG Image",
                                                     "div > div > div.home-kv-carousel__background-media-wrap",
                                                     bg_image_desktop_width, bg_image_desktop_height, bg_image_mobile_width, bg_image_mobile_height, "N", "Y",raw_data_meta)
            if 'cm-g-text-block-container' in content_comp_name:

                if content_comp_div.select('div > div.ho-g-showcase-card-tab'):
                    col_location = process_label_text(exl_ws, content_comp_div, "", "", "", "",
                                                            "section > div > div", "H", "text-block-container__headline", "Y", "N",raw_data_meta)
                    for content_comp_co02 in content_comp_div.select('div.showcase-card-tab'):

                        tab_no = 0
                        for comp_co02_tab in content_comp_co02.select('div > ul.tab__list > li'):
                            tab_no += 1
                            cell_area = process_label_text(exl_ws, comp_co02_tab, "", "", "", "",
                                                            "N", "button", "tab__item-title", "N", "N",raw_data_meta)

                            comp_co02_layout = content_comp_co02.select('div.showcase-card-tab__card-wrap > div > div.showcase-card-tab__card-items')[tab_no - 1]

                            card_layout_option = ''
                            card_no = 0

                            for card_compo in comp_co02_layout.find_all("div", class_='showcase-card-tab-card'):
                                if 'showcase-card-tab-card--large' in card_compo.attrs['class'] or 'showcase-card-tab-card--product-large' in card_compo.attrs['class'] or 'showcase-card-tab-card--vertical' in card_compo.attrs['class']:
                                    card_layout_option += 'L'
                                elif 'showcase-card-tab-card--small' in card_compo.attrs['class'] or 'showcase-card-tab-card--product-small' in card_compo.attrs['class']:
                                    card_layout_option += 'S'

                            options = {'SSSSSSSS': ['8 Card',
                                                    ['Top Left', 'Bottom Left', 'Top Center Left', 'Bottom Center Left',
                                                     'Top Center Right', 'Bottom Center Right', 'Top Right', 'Bottom Right']],
                                       'LSSSSSS': ['7 Card (1-2-2-2)',
                                                   ['Left', 'Top Center Left', 'Bottom Center Left', 'Top Center Right',
                                                    'Bottom Center Right', 'Top Right', 'Bottom Right']],
                                       'SSLSSSS': ['7 Card (2-1-2-2)',
                                                   ['Top Left', 'Bottom Left', 'Center Left', 'Top Center Right',
                                                    'Bottom Center Right', 'Top Right', 'Bottom Right']],
                                       'SSSSLSS': ['7 Card (2-2-1-2)',
                                                   ['Top Left', 'Bottom Left', 'Top Center Left', 'Bottom Center Left',
                                                    'Center Right', 'Top Right', 'Bottom Right']],
                                       'SSSSSSL': ['7 Card (2-2-2-1)',
                                                   ['Top Left', 'Bottom Left', 'Top Center Left', 'Bottom Center Left',
                                                    'Top Center Right', 'Bottom Center Right', 'Right']],
                                       'SSSSSS': ['6 Card',
                                                  ['Top Left', 'Bottom Left', 'Top Center', 'Bottom Center', 'Top Right',
                                                   'Bottom Right']],
                                       'SSSSL': ['5 Card (2-2-1)',
                                                 ['Top Left', 'Bottom Left', 'Top Center', 'Bottom Center', 'Right']],
                                       'SSLSS': ['5 Card (2-1-2)',
                                                 ['Top Left', 'Bottom Left', 'Center', 'Top Right', 'Bottom Right']],
                                       'LSSSS': ['5 Card (1-2-2)',
                                                 ['Left', 'Top Center', 'Bottom Center', 'Top Right', 'Bottom Right']],
                                       'LLL': ['3 Card', ['Left', 'Center', 'Right']]}

                            for x in options.keys():
                                if x == card_layout_option:
                                    card_number = options[x][0]
                                    card_layout = options[x][1]

                            card_list_no = 0
                            for card_list in card_layout:
                                card_list_no += 1

                                card_compo_detail = comp_co02_layout.find_all("div", class_='showcase-card-tab-card')[card_list_no - 1]
                                card_type = process_class_attributes(exl_ws, card_compo_detail, "", "", "", "",
                                                                      'N', 'N',
                                                                      {"showcase-card-tab-card--large": "Full Breed",
                                                                       "showcase-card-tab-card--small": "Full Breed",
                                                                       "showcase-card-tab-card--product-large": "Product",
                                                                       "showcase-card-tab-card--product-small": "Product"}, 'N',
                                                                      "N",raw_data_meta)

                                if card_type == 'Full Breed':
                                    if card_number == '3 Card':
                                        bg_image_desktop_width = 448
                                        bg_image_desktop_height = 684
                                        bg_image_mobile_width = 624
                                        bg_image_mobile_height = 352

                                    if (card_number == '5 Card (1-2-2)' and card_list == 'Left') or (
                                            card_number == '5 Card (2-1-2)' and card_list == 'Center') or (
                                            card_number == '5 Card (2-2-1)' and card_list == 'Right'):
                                        bg_image_desktop_width = 684
                                        bg_image_desktop_height = 684
                                        bg_image_mobile_width = 624
                                        bg_image_mobile_height = 352
                                    elif card_number == '5 Card (1-2-2)' or card_number == '5 Card (2-1-2)' or card_number == '5 Card (2-2-1)':
                                        bg_image_desktop_width = 330
                                        bg_image_desktop_height = 330
                                        bg_image_mobile_width = 296
                                        bg_image_mobile_height = 352

                                    if card_number == '6 Card':
                                        bg_image_desktop_width = 448
                                        bg_image_desktop_height = 330
                                        bg_image_mobile_width = 296
                                        bg_image_mobile_height = 352

                                    if (card_number == '7 Card (2-2-2-1)' and card_list == 'Right') or (
                                            card_number == '7 Card (2-2-1-2)' and card_list == 'Center Right') or (
                                            card_number == '7 Card (2-1-2-2)' and card_list == 'Center Left') or (
                                            card_number == '7 Card (1-2-2-2)' and card_list == 'Left'):
                                        bg_image_desktop_width = 330
                                        bg_image_desktop_height = 684
                                        bg_image_mobile_width = 624
                                        bg_image_mobile_height = 352
                                    elif card_number == '7 Card (2-2-2-1)' or card_number == '7 Card (2-2-1-2)' or card_number == '7 Card (2-1-2-2)' or card_number == '7 Card (1-2-2-2)':
                                        bg_image_desktop_width = 330
                                        bg_image_desktop_height = 330
                                        bg_image_mobile_width = 296
                                        bg_image_mobile_height = 352

                                    if card_number == '8 Card':
                                        bg_image_desktop_width = 330
                                        bg_image_desktop_height = 330
                                        bg_image_mobile_width = 296
                                        bg_image_mobile_height = 352

                                if card_type == 'Product':
                                    if card_number == '3 Card':
                                        bg_image_desktop_width = 376
                                        bg_image_desktop_height = 376
                                        bg_image_mobile_width = 264
                                        bg_image_mobile_height = 264

                                    if (card_number == '5 Card (1-2-2)' and card_list == 'Left') or (
                                            card_number == '5 Card (2-1-2)' and card_list == 'Center') or (
                                            card_number == '5 Card (2-2-1)' and card_list == 'Right'):
                                        bg_image_desktop_width = 376
                                        bg_image_desktop_height = 376
                                        bg_image_mobile_width = 264
                                        bg_image_mobile_height = 264
                                    elif card_number == '5 Card (1-2-2)' or card_number == '5 Card (2-1-2)' or card_number == '5 Card (2-2-1)':
                                        bg_image_desktop_width = 160
                                        bg_image_desktop_height = 160
                                        bg_image_mobile_width = 192
                                        bg_image_mobile_height = 192

                                    if card_number == '6 Card':
                                        bg_image_desktop_width = 160
                                        bg_image_desktop_height = 160
                                        bg_image_mobile_width = 192
                                        bg_image_mobile_height = 192

                                    if (card_number == '7 Card (2-2-2-1)' and card_list == 'Right') or (
                                            card_number == '7 Card (2-2-1-2)' and card_list == 'Center Right') or (
                                            card_number == '7 Card (2-1-2-2)' and card_list == 'Center Left') or (
                                            card_number == '7 Card (1-2-2-2)' and card_list == 'Left'):
                                        bg_image_desktop_width = 282
                                        bg_image_desktop_height = 282
                                        bg_image_mobile_width = 264
                                        bg_image_mobile_height = 264
                                    elif card_number == '7 Card (2-2-2-1)' or card_number == '7 Card (2-2-1-2)' or card_number == '7 Card (2-1-2-2)' or card_number == '7 Card (1-2-2-2)':
                                        bg_image_desktop_width = 160
                                        bg_image_desktop_height = 160
                                        bg_image_mobile_width = 192
                                        bg_image_mobile_height = 192

                                    if card_number == '8 Card':
                                        bg_image_desktop_width = 160
                                        bg_image_desktop_height = 160
                                        bg_image_mobile_width = 192
                                        bg_image_mobile_height = 192
                                if " " in card_list:
                                    card_size = "Small"
                                else:
                                    card_size = "Large"
                                cell_value = process_label_text(exl_ws, card_compo_detail, col_location, cell_area + " | " + card_list + "(" + card_size + ")", "Text", "Desktop",
                                                                      "N", "span",
                                                                      "showcase-card-tab-card__product-name--desktop", "Y", "Y",raw_data_meta)
                                cell_value = process_label_text(exl_ws, card_compo_detail, col_location, cell_area + " | " + card_list + "(" + card_size + ")", "Text", "Mobile",
                                                                "N", "span",
                                                                "showcase-card-tab-card__product-name--mobile", "Y", "Y",raw_data_meta)

                                cell_value = process_label_text(exl_ws, card_compo_detail, col_location,
                                                                cell_area + " | " + card_list + "(" + card_size + ")", "Text",
                                                                "Desktop",
                                                                "N", "span",
                                                                "showcase-card-tab-card__product-description--desktop", "N", "Y",raw_data_meta)
                                cell_value = process_label_text(exl_ws, card_compo_detail, col_location,
                                                                cell_area + " | " + card_list + "(" + card_size + ")", "Text",
                                                                "Mobile",
                                                                "N", "span",
                                                                "showcase-card-tab-card__product-description--mobile", "N", "Y",raw_data_meta)
                                process_label_text_cta(exl_ws, card_compo_detail, col_location, cell_area + " | " + card_list + "(" + card_size + ")", "CTA",
                                                                                     "N", "span.cta",raw_data_meta)

                                process_background_image(exl_ws, card_compo_detail, col_location, cell_area + " | " + card_list + "(" + card_size + ")",
                                                         "BG Image",
                                                         "div.showcase-card-tab-card__img-wrap",
                                                         bg_image_desktop_width, bg_image_desktop_height,
                                                         bg_image_mobile_width, bg_image_mobile_height, "Y", "Y",raw_data_meta)

            if 'ho-g-showcase-card-tab' in content_comp_name:

                col_location = 'Latest Deals'
                for content_comp_co02 in content_comp_div.select('div.showcase-card-tab'):

                    tab_no = 0
                    for comp_co02_tab in content_comp_co02.select('div > ul.tab__list > li'):
                        tab_no += 1
                        cell_area = process_label_text(exl_ws, comp_co02_tab, "", "", "", "",
                                                       "N", "button", "tab__item-title", "N", "N",raw_data_meta)

                        comp_co02_layout = content_comp_co02.select(
                            'div.showcase-card-tab__card-wrap > div > div.showcase-card-tab__card-items')[tab_no - 1]

                        card_layout_option = ''
                        card_no = 0

                        for card_compo in comp_co02_layout.find_all("div", class_='showcase-card-tab-card'):
                            if 'showcase-card-tab-card--large' in card_compo.attrs[
                                'class'] or 'showcase-card-tab-card--product-large' in card_compo.attrs['class']:
                                card_layout_option += 'L'
                            elif 'showcase-card-tab-card--small' in card_compo.attrs[
                                'class'] or 'showcase-card-tab-card--product-small' in card_compo.attrs['class']:
                                card_layout_option += 'S'

                        options = {'SSSSSSSS': ['8 Card',
                                                ['Top Left', 'Bottom Left', 'Top Center Left', 'Bottom Center Left',
                                                 'Top Center Right', 'Bottom Center Right', 'Top Right',
                                                 'Bottom Right']],
                                   'LSSSSSS': ['7 Card (1-2-2-2)',
                                               ['Left', 'Top Center Left', 'Bottom Center Left', 'Top Center Right',
                                                'Bottom Center Right', 'Top Right', 'Bottom Right']],
                                   'SSLSSSS': ['7 Card (2-1-2-2)',
                                               ['Top Left', 'Bottom Left', 'Center Left', 'Top Center Right',
                                                'Bottom Center Right', 'Top Right', 'Bottom Right']],
                                   'SSSSLSS': ['7 Card (2-2-1-2)',
                                               ['Top Left', 'Bottom Left', 'Top Center Left', 'Bottom Center Left',
                                                'Center Right', 'Top Right', 'Bottom Right']],
                                   'SSSSSSL': ['7 Card (2-2-2-1)',
                                               ['Top Left', 'Bottom Left', 'Top Center Left', 'Bottom Center Left',
                                                'Top Center Right', 'Bottom Center Right', 'Right']],
                                   'SSSSSS': ['6 Card',
                                              ['Top Left', 'Bottom Left', 'Top Center', 'Bottom Center', 'Top Right',
                                               'Bottom Right']],
                                   'SSSSL': ['5 Card (2-2-1)',
                                             ['Top Left', 'Bottom Left', 'Top Center', 'Bottom Center', 'Right']],
                                   'SSLSS': ['5 Card (2-1-2)',
                                             ['Top Left', 'Bottom Left', 'Center', 'Top Right', 'Bottom Right']],
                                   'LSSSS': ['5 Card (1-2-2)',
                                             ['Left', 'Top Center', 'Bottom Center', 'Top Right', 'Bottom Right']],
                                   'LLL': ['3 Card', ['Left', 'Center', 'Right']]}

                        for x in options.keys():
                            if x == card_layout_option:
                                card_number = options[x][0]
                                card_layout = options[x][1]

                        card_list_no = 0
                        for card_list in card_layout:
                            card_list_no += 1

                            card_compo_detail = comp_co02_layout.find_all("div", class_='showcase-card-tab-card')[
                                card_list_no - 1]
                            card_type = process_class_attributes(exl_ws, card_compo_detail, "", "", "", "",
                                                                 'N', 'N',
                                                                 {"showcase-card-tab-card--large": "Full Breed",
                                                                  "showcase-card-tab-card--small": "Full Breed",
                                                                  "showcase-card-tab-card--product-large": "Product",
                                                                  "showcase-card-tab-card--product-small": "Product"},
                                                                 'N',
                                                                 "N",raw_data_meta)

                            if card_type == 'Full Breed':
                                if card_number == '3 Card':
                                    bg_image_desktop_width = 448
                                    bg_image_desktop_height = 684
                                    bg_image_mobile_width = 624
                                    bg_image_mobile_height = 352

                                if (card_number == '5 Card (1-2-2)' and card_list == 'Left') or (
                                        card_number == '5 Card (2-1-2)' and card_list == 'Center') or (
                                        card_number == '5 Card (2-2-1)' and card_list == 'Right'):
                                    bg_image_desktop_width = 684
                                    bg_image_desktop_height = 684
                                    bg_image_mobile_width = 624
                                    bg_image_mobile_height = 352
                                elif card_number == '5 Card (1-2-2)' or card_number == '5 Card (2-1-2)' or card_number == '5 Card (2-2-1)':
                                    bg_image_desktop_width = 330
                                    bg_image_desktop_height = 330
                                    bg_image_mobile_width = 296
                                    bg_image_mobile_height = 352

                                if card_number == '6 Card':
                                    bg_image_desktop_width = 448
                                    bg_image_desktop_height = 330
                                    bg_image_mobile_width = 296
                                    bg_image_mobile_height = 352

                                if (card_number == '7 Card (2-2-2-1)' and card_list == 'Right') or (
                                        card_number == '7 Card (2-2-1-2)' and card_list == 'Center Right') or (
                                        card_number == '7 Card (2-1-2-2)' and card_list == 'Center Left') or (
                                        card_number == '7 Card (1-2-2-2)' and card_list == 'Left'):
                                    bg_image_desktop_width = 330
                                    bg_image_desktop_height = 684
                                    bg_image_mobile_width = 624
                                    bg_image_mobile_height = 352
                                elif card_number == '7 Card (2-2-2-1)' or card_number == '7 Card (2-2-1-2)' or card_number == '7 Card (2-1-2-2)' or card_number == '7 Card (1-2-2-2)':
                                    bg_image_desktop_width = 330
                                    bg_image_desktop_height = 330
                                    bg_image_mobile_width = 296
                                    bg_image_mobile_height = 352

                                if card_number == '8 Card':
                                    bg_image_desktop_width = 330
                                    bg_image_desktop_height = 330
                                    bg_image_mobile_width = 296
                                    bg_image_mobile_height = 352

                            if card_type == 'Product':
                                if card_number == '3 Card':
                                    bg_image_desktop_width = 376
                                    bg_image_desktop_height = 376
                                    bg_image_mobile_width = 264
                                    bg_image_mobile_height = 264

                                if (card_number == '5 Card (1-2-2)' and card_list == 'Left') or (
                                        card_number == '5 Card (2-1-2)' and card_list == 'Center') or (
                                        card_number == '5 Card (2-2-1)' and card_list == 'Right'):
                                    bg_image_desktop_width = 376
                                    bg_image_desktop_height = 376
                                    bg_image_mobile_width = 264
                                    bg_image_mobile_height = 264
                                elif card_number == '5 Card (1-2-2)' or card_number == '5 Card (2-1-2)' or card_number == '5 Card (2-2-1)':
                                    bg_image_desktop_width = 160
                                    bg_image_desktop_height = 160
                                    bg_image_mobile_width = 192
                                    bg_image_mobile_height = 192

                                if card_number == '6 Card':
                                    bg_image_desktop_width = 160
                                    bg_image_desktop_height = 160
                                    bg_image_mobile_width = 192
                                    bg_image_mobile_height = 192

                                if (card_number == '7 Card (2-2-2-1)' and card_list == 'Right') or (
                                        card_number == '7 Card (2-2-1-2)' and card_list == 'Center Right') or (
                                        card_number == '7 Card (2-1-2-2)' and card_list == 'Center Left') or (
                                        card_number == '7 Card (1-2-2-2)' and card_list == 'Left'):
                                    bg_image_desktop_width = 282
                                    bg_image_desktop_height = 282
                                    bg_image_mobile_width = 264
                                    bg_image_mobile_height = 264
                                elif card_number == '7 Card (2-2-2-1)' or card_number == '7 Card (2-2-1-2)' or card_number == '7 Card (2-1-2-2)' or card_number == '7 Card (1-2-2-2)':
                                    bg_image_desktop_width = 160
                                    bg_image_desktop_height = 160
                                    bg_image_mobile_width = 192
                                    bg_image_mobile_height = 192

                                if card_number == '8 Card':
                                    bg_image_desktop_width = 160
                                    bg_image_desktop_height = 160
                                    bg_image_mobile_width = 192
                                    bg_image_mobile_height = 192
                            if " " in card_list:
                                card_size = "Small"
                            else:
                                card_size = "Large"
                            cell_value = process_label_text(exl_ws, card_compo_detail, col_location,
                                                            cell_area + " | " + card_list + "(" + card_size + ")",
                                                            "Text", "Desktop",
                                                            "N", "span",
                                                            "showcase-card-tab-card__product-name--desktop", "Y", "Y",raw_data_meta)
                            cell_value = process_label_text(exl_ws, card_compo_detail, col_location,
                                                            cell_area + " | " + card_list + "(" + card_size + ")",
                                                            "Text", "Mobile",
                                                            "N", "span",
                                                            "showcase-card-tab-card__product-name--mobile", "Y", "Y",raw_data_meta)

                            cell_value = process_label_text(exl_ws, card_compo_detail, col_location,
                                                            cell_area + " | " + card_list + "(" + card_size + ")",
                                                            "Text",
                                                            "Desktop",
                                                            "N", "span",
                                                            "showcase-card-tab-card__product-description--desktop", "N",
                                                            "Y",raw_data_meta)
                            cell_value = process_label_text(exl_ws, card_compo_detail, col_location,
                                                            cell_area + " | " + card_list + "(" + card_size + ")",
                                                            "Text",
                                                            "Mobile",
                                                            "N", "span",
                                                            "showcase-card-tab-card__product-description--mobile", "N",
                                                            "Y",raw_data_meta)
                            process_label_text_cta(exl_ws, card_compo_detail, col_location,
                                                   cell_area + " | " + card_list + "(" + card_size + ")", "CTA",
                                                   "N", "span.cta",raw_data_meta)

                            process_background_image(exl_ws, card_compo_detail, col_location,
                                                     cell_area + " | " + card_list + "(" + card_size + ")",
                                                     "BG Image",
                                                     "div.showcase-card-tab-card__img-wrap",
                                                     bg_image_desktop_width, bg_image_desktop_height,
                                                     bg_image_mobile_width, bg_image_mobile_height, "Y", "Y",raw_data_meta)

            if 'ho-g-key-feature-tab' in content_comp_name:
                col_location = process_label_text(exl_ws, content_comp_div, "", "", "", "",
                                                "section > div.key-feature-tab__contents > div.key-feature-tab__header-wrap",
                                                  "H", "key-feature-tab__title", "N", "N",raw_data_meta)
                # print(col_location)
                comp_tab_no = 0
                for comp_co07_tab in content_comp_div.select('section > div.key-feature-tab__contents > div.key-feature-tab__header-wrap > div > ul > li'):
                    comp_co07_contents = content_comp_div.select('section > div.key-feature-tab__contents > div.key-feature-tab__container > div > div.key-feature-tab__slide')[comp_tab_no]
                    comp_tab_no += 1

                    col_area = process_label_text(exl_ws, comp_co07_tab, "" , "", "", "",
                                                    "N", "button", "tab__item-title", "N", "N",raw_data_meta)

                    cell_value = process_label_text(exl_ws, comp_co07_contents, col_location, col_area, "Headline Text", "All",
                                                          "div > div.key-feature-tab__inner-wrap > div.key-feature-tab__text-wrap",
                                                          "H", "key-feature-tab__headline", "Y", "Y",raw_data_meta)
                    cell_value = process_label_text(exl_ws, comp_co07_contents, col_location, col_area, "Description Text",
                                                    "All",
                                                    "div > div.key-feature-tab__inner-wrap > div.key-feature-tab__text-wrap",
                                                    "p", "key-feature-tab__desc", "N", "Y",raw_data_meta)

                    process_cta_buttons(exl_ws, comp_co07_contents, col_location, col_area, "CTA",
                                        "div > div.key-feature-tab__inner-wrap > div.key-feature-tab__cta-wrap",raw_data_meta)

                    bg_image_desktop_width = 1440
                    bg_image_desktop_height = 810
                    bg_image_mobile_width = 720
                    bg_image_mobile_height = 1280

                    process_background_image(exl_ws, comp_co07_contents, col_location, col_area, "BG image",
                                             'div > div.key-feature-tab__background-wrap > div', bg_image_desktop_width,
                                             bg_image_desktop_height, bg_image_mobile_width, bg_image_mobile_height, "Y", "Y",raw_data_meta)

        row_num = exl_ws.max_row + 1
        for i in range(1, 9):
            exl_ws.cell(row=row_num, column=i).border = Border(top=Side(style='thin', color='555555'))

        df = pd.DataFrame(data_list)
        file_name_base = url.replace('https://www.samsung.com/', '').replace('/', '-')
        file_name_base = file_name_base.strip('-')
        current_time = datetime.now().strftime("%Y%m%d%H%M%S")
        today = datetime.now().strftime("%Y%m%d")
        path = "./result/"+today
        os.makedirs(path, exist_ok=True)
        final_file_name = f"{path}/{file_name_base}_{current_time}.xlsx"
        exl_wb.save(final_file_name)
        end_time = datetime.now()
        time_elapsed = end_time - start_time
        print("FINISH", url, 'at : ', end_time.strftime("%Y/%m/%d %H:%M:%S"), "(", time_elapsed, ")")
        print(f"SAVE AT : {final_file_name}")


    except Exception as e:
        file_name_base = url.replace('https://www.samsung.com/', '').replace('/', '-')
        file_name_base = file_name_base.strip('-')
        today = datetime.now().strftime("%Y%m%d")
        path = "./result/" + today + "/error"
        os.makedirs(path, exist_ok=True)
        current_time = datetime.now().strftime("%Y%m%d%H%M%S")
        final_file_name = f"{path}/{file_name_base}_eroror_{current_time}.txt"

        end_time = datetime.now()
        time_elapsed = end_time - start_time

        f = open(final_file_name, "w")
        # f.write(str(time_elapsed),"\n TEST")
        error_log = f"{e}, \n ERROR :, {url}, at : {end_time} ({time_elapsed})"
        f.write(error_log)
        f.close()

        print("ERROR :", url, 'at : ', end_time.strftime("%Y/%m/%d %H:%M:%S"), "(", time_elapsed, ")")
        print(e)